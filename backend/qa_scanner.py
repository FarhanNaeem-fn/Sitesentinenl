"""
SiteSentinel — QA scan engine: HTML/XLSX report generators, scan implementation, and QAEngine class.
"""
from __future__ import annotations

import asyncio
import json
import os
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import BROWSER_WS, CONFIG, REPORTS_DIR, SEV, VIEWPORTS, log, _run_in_proactor
from job_manager import jdone, jerr, jlog, jobs
from models import QAScanRequest
from report_utils import _REPORT_CSS, _report_badge, _score_color
from core import (
    _open_browser, _tc, calculate_site_health, ScreenshotMarker,
    _resolve_url, _url_preflight, _crawl_site_for_assets,
    _check_locations_for_qa, _describe_console_message,
)
from supabase_client import db


def _run_qa_scan(jid: str, req: QAScanRequest):
    """Sync wrapper — delegates to ProactorEventLoop for Playwright."""
    _run_in_proactor(_run_qa_scan_impl(jid, req))

def _generate_qa_html_report(results: dict) -> str:
    from report_engine import ReportBuilder, _score_col, _esc, _sev_badge

    meta         = results.get("health_meta", {})
    score        = results.get("health_score", 0)
    grade        = meta.get("grade", "N/A")
    url          = results.get("url", "Unknown")
    ts           = results.get("timestamp", "Unknown")
    all_tcs      = results.get("all_test_cases", [])
    image_issues = results.get("image_issues", [])
    link_issues  = results.get("link_issues", [])
    pages_crawled = results.get("pages_crawled", [])
    n_pages  = len(pages_crawled) or results.get("pages_scanned", 1)
    passed   = sum(1 for t in all_tcs if t.get("Result") == "PASS")
    failed   = sum(1 for t in all_tcs if t.get("Result") == "FAIL")
    total_tc = len(all_tcs)
    n_img    = len(image_issues)
    n_lnk    = len(link_issues)

    rb = ReportBuilder(f"QA Audit Report — {url}", url, "QA Scan", ts)
    rb.set_score(score, "Quality")

    rb.add_kpi("Health Score",  f"{score}/100",           f"Grade: {grade}",          _score_col(score))
    rb.add_kpi("Pages Scanned", str(n_pages),             "internal pages crawled",   "#3B82F6")
    rb.add_kpi("Tests Run",     str(total_tc),            f"{passed} passed · {failed} failed", "#8B5CF6")
    rb.add_kpi("Image Issues",  str(n_img),               "missing alt text",         "#F59E0B" if n_img else "#22C55E")
    rb.add_kpi("Broken Links",  str(n_lnk),               "HTTP errors found",        "#EF4444" if n_lnk else "#22C55E")
    rb.add_kpi("Total Issues",  str(failed + n_img + n_lnk), "items to fix",          "#EF4444" if (failed + n_img + n_lnk) else "#22C55E")

    rb.add_charts([
        {"id":"qa_results","title":"Test Result Distribution","type":"donut",
         "labels":["Passed","Failed"],
         "values":[passed, failed],
         "colors":["#22C55E","#EF4444"]},
        {"id":"qa_issues","title":"Issues by Type","type":"bar",
         "labels":["Image Issues","Broken Links","Failed Tests"],
         "values":[n_img, n_lnk, failed],
         "label":"Count",
         "colors":["#F59E0B","#EF4444","#8B5CF6"]},
    ])

    # ── Image Issues section ──────────────────────────────────────────
    if image_issues:
        img_rows = ""
        for item in image_issues:
            img_url = item.get("image_url","")
            short   = img_url if len(img_url) <= 90 else img_url[:87] + "…"
            img_rows += f"""<tr>
  <td class="rpt-td-name">{_esc(item.get('page_name',''))}</td>
  <td><a class="rpt-link" href="{_esc(item.get('page_url',''))}" target="_blank">{_esc(item.get('page_url',''))}</a></td>
  <td><a class="rpt-link rpt-mono" href="{_esc(img_url)}" target="_blank">{_esc(short)}</a></td>
  <td><span class="rpt-badge rpt-badge-warn">{_esc(item.get('issue',''))}</span></td>
  <td class="rpt-td-dim">{_esc(item.get('description',''))}<br><em style="font-size:10px;color:#484F58">{_esc(item.get('fix',''))}</em></td>
</tr>"""
        rb.add_section("Image Issues", "\U0001f5bc️",
            f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>Page</th><th>Page URL</th><th>Image URL</th><th>Issue</th><th>Description &amp; Fix</th>
  </tr></thead><tbody>{img_rows}</tbody></table>
</div></div>""",
            subtitle=f"{n_img} images with missing or empty alt text across {len(set(i['page_url'] for i in image_issues))} page(s)",
            badge=str(n_img))
    else:
        rb.add_section("Image Issues", "\U0001f5bc️",
            '<div class="rpt-card"><div class="rpt-card-body"><p style="color:#22C55E;font-weight:600">✓ No image issues detected — all images have proper alt attributes.</p></div></div>',
            badge="0")

    # ── Broken Links section ──────────────────────────────────────────
    if link_issues:
        lnk_rows = ""
        for item in link_issues:
            status  = item.get("status","ERR")
            st_col  = "#EF4444" if str(status).startswith(("4","5","E")) else "#22C55E"
            lnk_rows += f"""<tr>
  <td class="rpt-td-name">{_esc(item.get('page_name',''))}</td>
  <td><a class="rpt-link" href="{_esc(item.get('page_url',''))}" target="_blank">{_esc(item.get('page_url',''))}</a></td>
  <td><a class="rpt-link rpt-mono" href="{_esc(item.get('link_url',''))}" target="_blank">{_esc(item.get('link_url','')[:90])}</a></td>
  <td class="rpt-td-dim">{_esc(item.get('link_text',''))}</td>
  <td style="font-weight:700;color:{st_col};font-family:monospace">{_esc(status)}</td>
  <td><span class="rpt-badge rpt-badge-fail">{_esc(item.get('issue',''))}</span></td>
  <td class="rpt-td-dim">{_esc(item.get('fix',''))}</td>
</tr>"""
        rb.add_section("Broken Links", "\U0001f517",
            f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>Page</th><th>Page URL</th><th>Broken URL</th><th>Link Text</th><th>Status</th><th>Issue</th><th>Fix</th>
  </tr></thead><tbody>{lnk_rows}</tbody></table>
</div></div>""",
            subtitle=f"{n_lnk} links returning HTTP errors across {len(set(i['page_url'] for i in link_issues))} page(s)",
            badge=str(n_lnk))
    else:
        rb.add_section("Broken Links", "\U0001f517",
            '<div class="rpt-card"><div class="rpt-card-body"><p style="color:#22C55E;font-weight:600">✓ No broken links detected — all links are accessible.</p></div></div>',
            badge="0")

    # ── QA Test Results ───────────────────────────────────────────────
    qa_rows = ""
    for tc in all_tcs:
        is_pass   = tc.get("Result") == "PASS"
        sev       = (tc.get("Severity") or "low").lower()
        sev_color = {"critical":"#EF4444","high":"#F59E0B","medium":"#3B82F6","low":"#8B949E"}.get(sev,"#8B949E")
        res_col   = "#22C55E" if is_pass else "#EF4444"
        detail    = _esc(tc.get("Detail") or "")
        fix       = _esc(tc.get("fix_hint") or "")
        fix_html  = f'<div style="font-size:10px;color:#484F58;margin-top:4px;padding:3px 6px;border-left:2px solid #30363D">{fix}</div>' if fix else ""
        missing   = tc.get("missing_alt_images",[])
        alt_html  = ""
        if missing:
            links = "".join(f'<li><a class="rpt-link rpt-mono" href="{_esc(s)}" target="_blank">{_esc(s[:80])}</a></li>' for s in missing[:10])
            alt_html = f'<div style="margin-top:6px;padding:6px 10px;background:rgba(88,166,255,.05);border-radius:4px;border:1px solid rgba(88,166,255,.15)"><strong>Images missing alt ({len(missing)}):</strong><ul style="margin:4px 0 0 16px">{links}</ul></div>'
        row_style = "border-left:3px solid #22C55E" if is_pass else "border-left:3px solid #EF4444"
        qa_rows += f"""<tr style="{row_style}">
  <td class="rpt-mono rpt-td-dim" style="font-size:10px">{_esc(tc.get('ID',''))}</td>
  <td><span style="display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;background:rgba(139,92,246,.1);color:#a78bfa;border:1px solid rgba(139,92,246,.25)">{_esc(tc.get('Category',''))}</span></td>
  <td>{_esc(tc.get('Test Name', tc.get('Test Case','')))}</td>
  <td style="font-weight:800;color:{res_col}">{_esc(tc.get('Result',''))}</td>
  <td><span style="color:{sev_color};font-size:11px;font-weight:700;font-family:monospace">{_esc(tc.get('Severity','LOW'))}</span></td>
  <td>{detail}{alt_html}{fix_html}</td>
</tr>"""
    categories = len(set(t.get("Category","") for t in all_tcs))
    rb.add_section("QA Test Results", "\U0001f52c",
        f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>ID</th><th>Category</th><th>Test Name</th><th style="width:70px">Result</th><th>Severity</th><th>Detail &amp; Fix</th>
  </tr></thead><tbody>{qa_rows}</tbody></table>
</div></div>""",
        subtitle=f"{passed} passed · {failed} issues across {categories} categories",
        badge=str(total_tc))

    # ── Recommendations from failures ────────────────────────────────
    recs = []
    fail_cats = {}
    for tc in all_tcs:
        if tc.get("Result") != "PASS":
            cat = tc.get("Category","General")
            fail_cats[cat] = fail_cats.get(cat,0) + 1
    for cat, count in sorted(fail_cats.items(), key=lambda x: -x[1])[:6]:
        recs.append({
            "title": f"Fix {count} failing {cat} test(s)",
            "priority": "quick_win" if count >= 3 else "medium",
            "description": f"{count} test(s) failed in the {cat} category. Review each failure for root cause.",
            "effort": "Low–Medium", "impact": "Medium"
        })
    if n_img:
        recs.append({"title": f"Add alt text to {n_img} image(s)", "priority": "quick_win",
            "description": "Missing alt attributes reduce accessibility and SEO. Add descriptive alt text to every image.",
            "effort": "Low", "impact": "High"})
    if n_lnk:
        recs.append({"title": f"Fix {n_lnk} broken link(s)", "priority": "quick_win",
            "description": "Broken links harm SEO and user experience. Remove or redirect each broken URL.",
            "effort": "Low", "impact": "High"})
    if recs:
        rb.add_recommendations(recs)

    rb.add_raw_data({"url": url, "health_score": score, "pages_crawled": n_pages,
                     "tests": total_tc, "passed": passed, "failed": failed,
                     "image_issues": n_img, "link_issues": n_lnk})
    return rb.build()


def _generate_browser_html_report(results: dict) -> str:
    """Generates a premium HTML report for Browser Automation scans."""
    url = results.get("url", "Unknown")
    ts = results.get("timestamp", "Unknown")
    status = results.get("status", 0)
    ms = results.get("ms", 0)

    checks = results.get("checks", [])
    passed_count = len([c for c in checks if c.get("ok")])
    total_count = len(checks)
    failed_count = total_count - passed_count
    pass_ratio = (passed_count / total_count * 100) if total_count > 0 else 0

    # Performance rating
    perf_cls = "pass" if ms < 1000 else "warn" if ms < 3000 else "fail"
    perf_lbl = "GOOD" if ms < 1000 else "FAIR" if ms < 3000 else "POOR"

    rows = ""
    for r in checks:
        res_cls = "pass" if r.get("ok") else "fail"
        rows += f"""
        <tr class="{res_cls}">
            <td><strong>{r.get('name', 'Check')}</strong></td>
            <td><span class="badge {res_cls}">{ 'PASS' if r.get('ok') else 'FAIL' }</span></td>
            <td>{r.get('detail', '—')}</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Browser Automation Report — {url}</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&family=JetBrains+Mono&display=swap" rel="stylesheet">
        <style>
            :root {{
                --bg: #0D1117; --card: #161B22; --border: #30363D;
                --text: #C9D1D9; --white: #FFFFFF; --dim: #8B949E;
                --pass: #238636; --fail: #DA3633; --warn: #D29922; --accent: #3B82F6;
            }}
            body {{ font-family: 'Inter', sans-serif; background: var(--bg); color: var(--text); margin: 0; padding: 40px; }}
            .container {{ max-width: 1000px; margin: 0 auto; }}
            .header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 40px; }}
            .header h1 {{ font-weight: 800; color: var(--white); margin: 0; font-size: 28px; }}
            .header p {{ color: var(--dim); margin: 5px 0 0; font-family: 'JetBrains Mono'; font-size: 14px; }}

            .dashboard {{ display: grid; grid-template-columns: 300px 1fr; gap: 30px; margin-bottom: 40px; }}
            .chart-box {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; padding: 30px; text-align: center; display: flex; flex-direction: column; align-items: center; justify-content: center; }}

            .ring-chart {{ position: relative; width: 150px; height: 150px; margin-bottom: 20px; }}
            .ring-chart svg {{ transform: rotate(-90deg); }}
            .ring-chart .bg {{ fill: none; stroke: var(--border); stroke-width: 12; }}
            .ring-chart .perc {{ fill: none; stroke: var(--pass); stroke-width: 12; stroke-linecap: round; stroke-dasharray: 440; stroke-dashoffset: {440 - (440 * pass_ratio / 100)}; transition: stroke-dashoffset 1s ease-out; }}
            .ring-chart .label {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); font-size: 24px; font-weight: 800; color: var(--white); }}

            .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 20px; }}
            .stat-card {{ background: var(--card); border: 1px solid var(--border); padding: 25px; border-radius: 16px; transition: transform 0.2s; }}
            .stat-card:hover {{ transform: translateY(-2px); border-color: var(--dim); }}
            .stat-card label {{ display: block; font-size: 11px; text-transform: uppercase; color: var(--dim); margin-bottom: 10px; letter-spacing: 1.5px; font-weight: 600; }}
            .stat-card .val {{ font-size: 28px; font-weight: 800; color: var(--white); display: flex; align-items: center; gap: 10px; }}

            .speed-dot {{ width: 12px; height: 12px; border-radius: 50%; }}
            .speed-dot.pass {{ background: var(--pass); box-shadow: 0 0 10px var(--pass); }}
            .speed-dot.warn {{ background: var(--warn); box-shadow: 0 0 10px var(--warn); }}
            .speed-dot.fail {{ background: var(--fail); box-shadow: 0 0 10px var(--fail); }}

            table {{ width: 100%; border-collapse: separate; border-spacing: 0; background: var(--card); border: 1px solid var(--border); border-radius: 16px; overflow: hidden; margin-top: 20px; }}
            th {{ text-align: left; background: #21262D; padding: 20px; font-size: 12px; text-transform: uppercase; color: var(--dim); letter-spacing: 1px; }}
            td {{ padding: 20px; border-top: 1px solid var(--border); font-size: 14px; }}

            .tr-pass {{ border-left: 4px solid var(--pass); }}
            .tr-fail {{ border-left: 4px solid var(--fail); }}

            .badge {{ padding: 6px 12px; border-radius: 6px; font-size: 11px; font-weight: 800; font-family: 'JetBrains Mono'; text-transform: uppercase; }}
            .badge.pass {{ background: rgba(35,134,54,0.15); color: #3FB950; border: 1px solid rgba(35,134,54,0.3); }}
            .badge.fail {{ background: rgba(218,54,51,0.15); color: #FF7B72; border: 1px solid rgba(218,54,51,0.3); }}

            .trend-up {{ color: var(--pass); }}
            .trend-down {{ color: var(--fail); }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div>
                    <h1>Browser Automation</h1>
                    <p>Target: {url} • Generated: {ts}</p>
                </div>
                <div class="badge pass">SECURE SCAN</div>
            </div>

            <div class="dashboard">
                <div class="chart-box">
                    <div class="ring-chart">
                        <svg width="150" height="150">
                            <circle class="bg" cx="75" cy="75" r="70"></circle>
                            <circle class="perc" cx="75" cy="75" r="70"></circle>
                        </svg>
                        <div class="label">{int(pass_ratio)}%</div>
                    </div>
                    <div style="font-weight: 600; color: var(--white)">Success Rate</div>
                    <div style="font-size: 12px; color: var(--dim); margin-top: 5px">{passed_count} of {total_count} Passed</div>
                </div>

                <div class="stats-grid">
                    <div class="stat-card">
                        <label>Load Performance</label>
                        <div class="val">
                            <div class="speed-dot {perf_cls}"></div>
                            {perf_lbl}
                        </div>
                        <div style="font-size: 12px; color: var(--dim); margin-top: 5px">{ms}ms server response</div>
                    </div>
                    <div class="stat-card">
                        <label>Automation Status</label>
                        <div class="val" style="color: var(--white)">HTTP {status}</div>
                        <div style="font-size: 12px; color: var(--dim); margin-top: 5px">Connection established</div>
                    </div>
                    <div class="stat-card" style="grid-column: span 2">
                        <label>Scan Summary</label>
                        <div class="val" style="gap: 40px">
                            <div><span style="color: var(--pass)">{passed_count}</span> <span style="font-size: 14px; color: var(--dim)">PASSED</span></div>
                            <div><span style="color: var(--fail)">{failed_count}</span> <span style="font-size: 14px; color: var(--dim)">FAILED</span></div>
                            <div><span style="color: var(--white)">{total_count}</span> <span style="font-size: 14px; color: var(--dim)">TOTAL</span></div>
                        </div>
                    </div>
                </div>
            </div>

            <table>
                <thead><tr><th>Check Description</th><th>Result</th><th>Metric / Evidence</th></tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </div>
    </body>
    </html>
    """
    return html

def _generate_qa_xlsx_report(results: dict) -> bytes:
    """
    Generate a clean, professional Excel (.xlsx) QA report.
    No cell coloring — black text on white, bold headers on light gray.

    Sheets:
      1. Scan Summary    — metadata and key metrics
      2. Image Issues    — every image with alt/accessibility problems
      3. Broken Links    — every link that returned an HTTP error
      4. Issues Found    — every failing QA test case
      5. All Test Cases  — complete pass/fail matrix
    """
    try:
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — cannot generate XLSX report. "
                    "Run: pip install openpyxl>=3.1.0")
        return b""

    # ── Fix-hint lookup ───────────────────────────────────────────────
    _FIX_HINTS: dict = {
        "SEO-01":   "Add a <title> tag inside <head> with a descriptive page title (50-60 chars).",
        "SEO-03":   "Add <meta name='description' content='...'> in <head> (150-160 chars).",
        "SEO-05":   "Ensure exactly one <h1> heading per page to define the main topic.",
        "SEO-12":   "Add <meta name='viewport' content='width=device-width, initial-scale=1'> in <head>.",
        "SEO-14":   "Add descriptive alt attributes to every <img>: <img src='...' alt='Description'>.",
        "A11Y-01":  "Add lang attribute to <html>: <html lang='en'>.",
        "A11Y-02":  "Add descriptive alt text to all images: <img alt='Description of image'>.",
        "A11Y-03":  "Associate a <label for='id'> or aria-label with every form input.",
        "A11Y-04":  "Add a skip link as the first element: <a href='#main'>Skip to main content</a>.",
        "PERF-01":  "Reduce server response time: use caching, CDN, or upgrade server resources.",
        "PERF-02":  "Eliminate render-blocking: inline critical CSS, defer or async non-critical JS.",
        "PERF-04":  "Minimise total load time: compress images, minify JS/CSS, enable lazy loading.",
        "PERF-05":  "Reduce page weight: compress assets, use WebP images, enable gzip/brotli on server.",
        "SEC-01":   "Enable HTTPS with a valid SSL certificate; redirect all HTTP traffic to HTTPS.",
        "SEC-02":   "Add response header: Strict-Transport-Security: max-age=31536000; includeSubDomains",
        "SEC-03":   "Add response header: X-Content-Type-Options: nosniff",
        "SEC-05":   "Implement a Content-Security-Policy header to restrict which origins may load resources.",
        "IMG-01":   "Add alt attributes to all images: <img src='...' alt='A clear description'>.",
        "LINK-ERR": "Fix or remove broken links; ensure all linked URLs return HTTP 200.",
        "HTML-01":  "Add <!DOCTYPE html> as the very first line of the HTML document.",
        "HTML-02":  "Add <meta charset='UTF-8'> inside <head>.",
        "HTML-03":  "Add a favicon: <link rel='icon' href='/favicon.ico'>.",
        "NAV-01":   "Wrap primary navigation in a <nav> element with aria-label='Main navigation'>.",
        "NAV-04":   "Add a <footer> element containing contact info, legal links, and secondary nav.",
        "CONT-01":  "Add substantive content — at least 300 words to establish relevance for search engines.",
        "CONT-02":  "Remove placeholder 'Lorem ipsum' text and replace it with real content.",
        "TYPO-01":  "Set base font-size to at least 14px in CSS for comfortable readability.",
        "KEY-01":   "Ensure all interactive elements are keyboard-focusable (no tabindex='-1' on buttons/links).",
        "KEY-03":   "Add a visible skip-nav link as the first focusable element on every page.",
        "FORM-":    "Add CSRF protection token to all forms to prevent cross-site request forgery.",
        "MIXED-":   "Serve all resources over HTTPS; change any http:// resource URLs to https://.",
        "DOM-":     "Check domain configuration: DNS, SSL expiry, robots.txt, and sitemap.xml.",
    }

    def _get_fix(tc: dict) -> str:
        """Return fix hint: explicit > ID-prefix lookup > generic by category."""
        explicit = (tc.get("fix_hint") or "").strip()
        if explicit:
            return explicit
        tc_id = tc.get("ID", "")
        for prefix, hint in _FIX_HINTS.items():
            if tc_id.startswith(prefix):
                return hint
        cat = tc.get("Category", "").lower()
        defaults = {
            "seo":           "Review SEO meta tags, headings structure, and alt attributes.",
            "accessibility": "Audit against WCAG 2.1 AA: alt text, labels, colour contrast, and keyboard access.",
            "performance":   "Profile the page with Lighthouse or WebPageTest and address the slowest resources.",
            "security":      "Add missing HTTP security headers and ensure HTTPS is enforced.",
            "links":         "Crawl the site for broken links and fix or remove them.",
            "images":        "Add alt attributes to all images and compress large image files.",
            "console":       "Open browser DevTools console, identify errors, and fix the root cause.",
            "content":       "Replace placeholder text with real, keyword-rich content.",
        }
        for key, hint in defaults.items():
            if key in cat:
                return hint
        return "Review the issue detail and consult the relevant web standard or best-practice guide."

    def _get_affected_element(tc: dict) -> str:
        """Return the specific element affected: image src list for alt issues, else detail/actual."""
        imgs = tc.get("missing_alt_images")
        if imgs:
            return " | ".join(str(s) for s in imgs[:5]) + (" …" if len(imgs) > 5 else "")
        actual = (tc.get("Actual") or "").strip()
        if actual and actual not in ("", "N/A"):
            return actual
        detail = (tc.get("Detail") or "").strip()
        # keep it short in this column
        return detail[:120] if detail else ""

    try:
        wb = Workbook()
        url     = str(results.get("url", "Unknown"))
        ts      = str(results.get("timestamp", "Unknown"))
        score   = results.get("health_score", 0)
        grade   = (results.get("health_meta") or {}).get("grade", "N/A")
        all_tcs = [t for t in (results.get("all_test_cases") or []) if isinstance(t, dict)]
        passed  = sum(1 for t in all_tcs if t.get("Result") == "PASS")
        failed  = sum(1 for t in all_tcs if t.get("Result") == "FAIL")
        issues  = [t for t in all_tcs if t.get("Result") == "FAIL"]

        # ── Clean style constants (no color — professional client-ready) ─
        # Fonts
        TITLE_F  = Font(name="Calibri", bold=True,  color="000000", size=14)
        HDR_F    = Font(name="Calibri", bold=True,  color="000000", size=10)
        BODY_F   = Font(name="Calibri",              color="000000", size=10)
        LABEL_F  = Font(name="Calibri", bold=True,  color="000000", size=10)
        MONO_F   = Font(name="Courier New",          color="000000", size=9)
        PASS_F   = Font(name="Calibri", bold=True,  color="166534", size=10)  # dark green text only
        FAIL_F   = Font(name="Calibri", bold=True,  color="991B1B", size=10)  # dark red text only

        # Fills (white only — no backgrounds)
        WHITE    = PatternFill(fill_type="solid", fgColor="FFFFFF")
        HDR_FILL = PatternFill(fill_type="solid", fgColor="F0F0F0")  # light gray for headers

        # Borders
        THIN   = Side(style="thin",   color="BBBBBB")
        MED    = Side(style="medium", color="999999")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        HDR_BDR = Border(left=MED, right=MED, top=MED, bottom=MED)

        # Alignments
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
        LCENT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        def _hdr(ws, row, col, value, width=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = HDR_F; c.fill = HDR_FILL; c.alignment = CENTER; c.border = HDR_BDR
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return c

        def _cell(ws, row, col, value, font=None, align=None):
            c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
            c.font      = font  or BODY_F
            c.fill      = WHITE
            c.alignment = align or LEFT
            c.border    = BORDER
            return c

        image_issues = results.get("image_issues", [])
        link_issues  = results.get("link_issues",  [])
        n_img = len(image_issues)
        n_lnk = len(link_issues)
        pages_crawled = results.get("pages_crawled", [])
        n_pages = len(pages_crawled) or results.get("pages_scanned", 1)

        # ════════════════════════════════════════════════════════════
        # Sheet 1 — Scan Summary
        # ════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Scan Summary"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 62

        # Title row
        ws1.merge_cells("A1:B1")
        c = ws1["A1"]
        c.value = "SiteSentinel — Website QA Audit Report"
        c.font = TITLE_F; c.fill = WHITE; c.alignment = CENTER
        c.border = HDR_BDR
        ws1.row_dimensions[1].height = 32

        # Blank separator row
        ws1.merge_cells("A2:B2")
        ws1["A2"].fill = WHITE

        summary_rows = [
            ("Website URL",        url),
            ("Scan Date / Time",   ts),
            ("Health Score",       f"{score} / 100"),
            ("Grade",              grade),
            ("Pages Crawled",      str(n_pages)),
            ("Total QA Checks",    str(len(all_tcs))),
            ("Checks Passed",      str(passed)),
            ("Checks Failed",      str(failed)),
            ("Image Issues Found", str(n_img)),
            ("Broken Links Found", str(n_lnk)),
        ]
        for ri, (lbl, val) in enumerate(summary_rows, start=3):
            la = ws1.cell(row=ri, column=1, value=lbl)
            la.font = LABEL_F; la.fill = HDR_FILL; la.alignment = LCENT; la.border = BORDER
            va = ws1.cell(row=ri, column=2, value=val)
            fnt = PASS_F if lbl == "Checks Passed" else FAIL_F if lbl in ("Checks Failed","Image Issues Found","Broken Links Found") and int(val or 0) > 0 else BODY_F
            va.font = fnt; va.fill = WHITE; va.alignment = LCENT; va.border = BORDER
            ws1.row_dimensions[ri].height = 20

        # ════════════════════════════════════════════════════════════
        # Sheet 2 — Image Issues
        # ════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Image Issues")
        ws2.sheet_view.showGridLines = False

        img_cols = [
            ("Page Name",          22),
            ("Page URL",           48),
            ("Image URL",          62),
            ("Issue",              30),
            ("Description",        52),
            ("Recommended Fix",    50),
        ]
        for ci, (h, w) in enumerate(img_cols, start=1):
            _hdr(ws2, 1, ci, h, width=w)
        ws2.row_dimensions[1].height = 26

        if not image_issues:
            ws2.merge_cells(f"A2:{get_column_letter(len(img_cols))}2")
            c = ws2.cell(row=2, column=1, value="No image issues found — all images have correct alt attributes.")
            c.font = PASS_F; c.fill = WHITE; c.alignment = CENTER; c.border = BORDER
            ws2.row_dimensions[2].height = 24
        else:
            for ri, item in enumerate(image_issues, start=2):
                vals = [
                    item.get("page_name", ""),
                    item.get("page_url", ""),
                    item.get("image_url", ""),
                    item.get("issue", ""),
                    item.get("description", ""),
                    item.get("fix", ""),
                ]
                for ci, val in enumerate(vals, start=1):
                    fnt = MONO_F if ci == 3 else (FAIL_F if ci == 4 else BODY_F)
                    _cell(ws2, ri, ci, val, font=fnt)
                ws2.row_dimensions[ri].height = 34

        # ════════════════════════════════════════════════════════════
        # Sheet 3 — Broken Links
        # ════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Broken Links")
        ws3.sheet_view.showGridLines = False

        lnk_cols = [
            ("Page Name",         22),
            ("Page URL",          48),
            ("Broken Link URL",   62),
            ("Link Text",         28),
            ("HTTP Status",       14),
            ("Issue",             30),
            ("Recommended Fix",   50),
        ]
        for ci, (h, w) in enumerate(lnk_cols, start=1):
            _hdr(ws3, 1, ci, h, width=w)
        ws3.row_dimensions[1].height = 26

        if not link_issues:
            ws3.merge_cells(f"A2:{get_column_letter(len(lnk_cols))}2")
            c = ws3.cell(row=2, column=1, value="No broken links found — all checked links are accessible.")
            c.font = PASS_F; c.fill = WHITE; c.alignment = CENTER; c.border = BORDER
            ws3.row_dimensions[2].height = 24
        else:
            for ri, item in enumerate(link_issues, start=2):
                vals = [
                    item.get("page_name", ""),
                    item.get("page_url", ""),
                    item.get("link_url", ""),
                    item.get("link_text", ""),
                    item.get("status", ""),
                    item.get("issue", ""),
                    item.get("fix", ""),
                ]
                for ci, val in enumerate(vals, start=1):
                    fnt = MONO_F if ci in (2, 3) else (FAIL_F if ci == 5 else BODY_F)
                    _cell(ws3, ri, ci, val, font=fnt)
                ws3.row_dimensions[ri].height = 34

        # ════════════════════════════════════════════════════════════
        # Sheet 4 — Issues Found  (failed QA test cases)
        # ════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Issues Found")
        ws4.sheet_view.showGridLines = False

        issues = [t for t in all_tcs if t.get("Result") == "FAIL"]
        qa_issue_cols = [
            ("Category",              22),
            ("Test Name",             34),
            ("Page URL",              48),
            ("Issue Description",     52),
            ("Severity",              14),
            ("Recommended Fix",       52),
        ]
        for ci, (h, w) in enumerate(qa_issue_cols, start=1):
            _hdr(ws4, 1, ci, h, width=w)
        ws4.row_dimensions[1].height = 26

        if not issues:
            ws4.merge_cells(f"A2:{get_column_letter(len(qa_issue_cols))}2")
            c = ws4.cell(row=2, column=1, value="No QA issues found — all checks passed.")
            c.font = PASS_F; c.fill = WHITE; c.alignment = CENTER; c.border = BORDER
            ws4.row_dimensions[2].height = 24
        else:
            for ri, tc in enumerate(issues, start=2):
                vals = [
                    tc.get("Category", ""),
                    tc.get("Test Name", tc.get("Test Case", "")),
                    tc.get("page_url", url),
                    tc.get("Detail", "") or tc.get("Test Name", ""),
                    tc.get("Severity", "LOW"),
                    _get_fix(tc),
                ]
                for ci, val in enumerate(vals, start=1):
                    fnt = FAIL_F if ci == 5 else BODY_F
                    _cell(ws4, ri, ci, val, font=fnt)
                ws4.row_dimensions[ri].height = 34

        # ════════════════════════════════════════════════════════════
        # Sheet 5 — All Test Cases
        # ════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("All Test Cases")
        ws5.sheet_view.showGridLines = False

        all_cols = [
            ("ID",               13),
            ("Category",         22),
            ("Test Name",        38),
            ("Result",           10),
            ("Severity",         14),
            ("Detail",           50),
            ("Recommended Fix",  52),
        ]
        for ci, (h, w) in enumerate(all_cols, start=1):
            _hdr(ws5, 1, ci, h, width=w)
        ws5.row_dimensions[1].height = 26

        for ri, tc in enumerate(all_tcs, start=2):
            res = tc.get("Result", "")
            vals = [
                tc.get("ID", ""),
                tc.get("Category", ""),
                tc.get("Test Name", tc.get("Test Case", "")),
                res,
                tc.get("Severity", "LOW"),
                tc.get("Detail", ""),
                _get_fix(tc),
            ]
            for ci, val in enumerate(vals, start=1):
                fnt = PASS_F if (ci == 4 and res == "PASS") else FAIL_F if (ci == 4 and res == "FAIL") else BODY_F
                _cell(ws5, ri, ci, val, font=fnt)
            ws5.row_dimensions[ri].height = 28

        # ── Save ─────────────────────────────────────────────────────
        buf = _io.BytesIO()
        wb.save(buf)
        log.info(f"XLSX report generated: {buf.tell()} bytes — "
                 f"{len(issues)} QA issues, {n_img} image issues, {n_lnk} broken links")
        return buf.getvalue()

    except Exception as exc:
        log.exception(f"_generate_qa_xlsx_report failed: {exc}")
        return b""


async def _run_qa_scan_impl(jid: str, req: QAScanRequest):
    try:
        from playwright.async_api import async_playwright
        jlog(jid, "="*52, "hdr")
        jlog(jid, f"  QA SCAN  —  {req.url}", "hdr")
        jlog(jid, f"  Viewport: {req.viewport}  Max pages: {req.max_pages}", "hdr")
        jlog(jid, "="*52, "hdr")

        # ── Pre-check: resolve URL and confirm accessibility ─────────
        jlog(jid, "Checking URL accessibility...", "info")
        ok, final_url, status = await _resolve_url(req.url)
        if not ok:
            jerr(jid, f"URL not accessible ({status}): {req.url}")
            return
        if final_url != req.url:
            jlog(jid, f"URL resolved: {req.url} → {final_url}", "ok")
            req = req.model_copy(update={"url": final_url})

        VPs = {"desktop":{"width":1920,"height":1080},"mac":{"width":1440,"height":900},
               "laptop":{"width":1366,"height":768},"mobile":{"width":430,"height":932}}
        vp = VPs.get(req.viewport, VPs["desktop"])
        results = {"url":req.url,"viewport":req.viewport,"pages_scanned":0,
                   "total_issues":0,"checks_passed":0,"checks_failed":0,
                   "health_score":0,"details":{},"domain_health":{},
                   "timestamp":datetime.now().isoformat()}

        async with async_playwright() as pw:
            if BROWSER_WS:
                jlog(jid, "Connecting to remote browser...", "info")
                try:
                    browser = await asyncio.wait_for(
                        pw.chromium.connect_over_cdp(BROWSER_WS),
                        timeout=10
                    )
                    jlog(jid, "✓ Remote browser connected", "ok")
                except asyncio.TimeoutError:
                    jlog(jid, "Remote browser connection timeout, falling back to local", "warn")
                    browser = await pw.chromium.launch(headless=True)
                except Exception as e:
                    jlog(jid, f"Remote browser connection failed ({e}), falling back to local", "warn")
                    browser = await pw.chromium.launch(headless=True)
            else:
                jlog(jid, "Launching local Chromium...", "info")
                browser = await pw.chromium.launch(headless=True)

            try:
                ctx = await browser.new_context(viewport=vp)
                page = await ctx.new_page()
                jlog(jid, "✓ Browser context and page created", "ok")
            except Exception as e:
                jlog(jid, f"✗ Failed to create context/page: {e}", "err")
                jerr(jid, f"Browser setup failed: {str(e)}")
                try:
                    await browser.close()
                except:
                    pass
                return

            console_errors = []
            page.on("console", lambda m: console_errors.append(m.text)
                    if m.type=="error" else None)

            # Validate browser connection is responsive (especially important on Vercel)
            try:
                jlog(jid, "Validating browser connection...", "info")
                await asyncio.wait_for(page.evaluate("1+1"), timeout=5)
                jlog(jid, "✓ Browser connection validated", "ok")
            except Exception as e:
                jlog(jid, f"✗ Browser validation failed: {str(e)[:60]}", "err")
                jerr(jid, "Browser connection failed validation")
                try:
                    await browser.close()
                except:
                    pass
                return

            # Configure navigation timeout and try robust navigation strategy
            try:
                nav_timeout = int(CONFIG.get("SCAN_GOTO_TIMEOUT", 60000))
            except Exception:
                nav_timeout = 60000
            try:
                await page.set_default_navigation_timeout(nav_timeout)
            except Exception:
                pass

            nav_success = False
            try:
                # First attempt: networkidle with reduced timeout to avoid remote session timeout on Vercel
                initial_timeout = min(nav_timeout, 45000)  # Cap at 45s to avoid Vercel remote service timeout
                jlog(jid, f"Navigating to {req.url} (timeout: {initial_timeout}ms, wait: networkidle)...", "info")
                resp = await page.goto(req.url, wait_until="networkidle", timeout=initial_timeout)
                results["pages_scanned"] = 1
                sc = resp.status if resp else 0
                jlog(jid, f"✓ Page loaded — HTTP {sc}", "ok")
                nav_success = True
            except Exception as e:
                jlog(jid, f"Navigation (networkidle) failed: {str(e)[:80]}", "warn")

                # Check if page/context/browser still alive before retry
                try:
                    test_title = await asyncio.wait_for(page.title(), timeout=2)
                    jlog(jid, "Page still responsive, attempting fallback", "info")
                except Exception as check_err:
                    jlog(jid, f"Page unresponsive after first attempt: {str(check_err)[:60]}", "err")
                    # Page is likely closed; recreate it
                    try:
                        jlog(jid, "Recreating page for second attempt", "info")
                        page = await ctx.new_page()
                        page.on("console", lambda m: console_errors.append(m.text)
                                if m.type=="error" else None)
                    except Exception as recreate_err:
                        jlog(jid, f"Failed to recreate page: {str(recreate_err)[:60]}", "err")
                        jerr(jid, f"Navigation failed: Page recreation failed")
                        try:
                            await browser.close()
                        except:
                            pass
                        return

                # Fallback: try with 'load' which is less strict for pages with ongoing network
                try:
                    fallback_timeout = min(nav_timeout*2, 90000)  # Cap at 90s
                    jlog(jid, f"Retrying with wait: load (timeout: {fallback_timeout}ms)...", "info")
                    resp = await page.goto(req.url, wait_until="load", timeout=fallback_timeout)
                    results["pages_scanned"] = 1
                    sc = resp.status if resp else 0
                    jlog(jid, f"✓ Page loaded (fallback load) — HTTP {sc}", "ok")
                    nav_success = True
                except Exception as e2:
                    jlog(jid, f"✗ Navigation failed (both strategies): {str(e2)[:80]}", "err")
                    jerr(jid, f"Navigation failed: {str(e2)[:200]}")
                    try:
                        await browser.close()
                    except:
                        pass
                    return

            if not nav_success:
                jerr(jid, "Navigation failed: Page did not load")
                try:
                    await browser.close()
                except:
                    pass
                return

            tp=0; tf=0; n=len(req.checks)
            all_test_cases = []
            results["all_test_cases"] = []

            for step_i, chk in enumerate(req.checks):
                jobs[jid]["progress"] = int((step_i/n)*80)+5
                r = {"test_cases": []}

                if chk=="seo":
                    jlog(jid, "Running: SEO Analysis", "info")
                    r = await QAEngine.check_seo(page)
                elif chk=="accessibility":
                    jlog(jid, "Running: Accessibility (WCAG 2.1 AA)", "info")
                    r = await QAEngine.check_accessibility(page)
                elif chk=="performance":
                    jlog(jid, "Running: Performance Metrics", "info")
                    r = await QAEngine.check_performance(page)
                elif chk=="security":
                    jlog(jid, "Running: Security Headers", "info")
                    r = await QAEngine.check_security(page, resp)
                elif chk=="console":
                    jlog(jid, "Running: Console Errors", "info")
                    r = await QAEngine.build_console_test_cases(console_errors)
                elif chk=="responsive":
                    jlog(jid, "Running: Responsive Layout (4 viewports)", "info")
                    r = await QAEngine.check_responsive(page, ctx, req.url)
                elif chk=="mixed_content":
                    jlog(jid, "Running: Mixed Content Check", "info")
                    r = await QAEngine.check_mixed_content(page)
                elif chk=="content":
                    jlog(jid, "Running: Content Quality Audit", "info")
                    r = await QAEngine.check_content_quality(page)
                elif chk=="broken_links":
                    jlog(jid, "Running: Broken Links", "info")
                    r = await QAEngine.check_broken_links(page, req.url)
                elif chk=="keyboard_access":
                    jlog(jid, "Running: Keyboard Accessibility", "info")
                    r = await QAEngine.check_keyboard_access(page)
                elif chk=="forms":
                    jlog(jid, "Running: Form Security", "info")
                    r = await QAEngine.check_forms(page)
                elif chk=="typography":
                    jlog(jid, "Running: Typography Analysis", "info")
                    r = await QAEngine.check_typography(page)
                elif chk=="html_quality":
                    jlog(jid, "Running: HTML Quality", "info")
                    r = await QAEngine.check_html_quality(page)
                elif chk=="images":
                    jlog(jid, "Running: Image Optimization", "info")
                    r = await QAEngine.check_images(page)
                elif chk=="navigation":
                    jlog(jid, "Running: Navigation/UX", "info")
                    r = await QAEngine.check_navigation(page)
                elif chk=="health_score":
                    jlog(jid, "Running: Site Health Score", "info")
                    r = {"test_cases": [], "note": "Calculated at end"}

                # Simplified check for other legacy types
                results["details"][chk] = r
                tcs = r.get("test_cases", [])
                all_test_cases.extend(tcs)
                results["all_test_cases"].extend(tcs)

                p = sum(1 for t in tcs if t["Result"] == "PASS")
                f = sum(1 for t in tcs if t["Result"] == "FAIL")
                tp += p; tf += f
                lv = "ok" if f==0 else "warn" if f<3 else "err"
                jlog(jid, f"  {chk}: {p} passed / {f} failed", lv)

                jobs[jid]["partial"] = {
                    "current_check": chk, "checks_done": step_i+1, "checks_total": n,
                    "checks_passed": tp, "checks_failed": tf,
                    "details": {k: {"passed": sum(1 for t in v.get("test_cases",[]) if t["Result"]=="PASS"),
                                    "failed": sum(1 for t in v.get("test_cases",[]) if t["Result"]=="FAIL")}
                                for k, v in results["details"].items()},
                }

            # Domain / site health
            if "health_score" in req.checks:
                health = await _run_domain_health(jid, req.url, page)
                results["domain_health"] = health
                # Add domain health test cases to the list
                tcs_domain = []
                for k, v in health.get("checks", {}).items():
                    res = "PASS" if v.get("ok") else "FAIL"
                    tc_item = _tc(f"DOM-{k[:3].upper()}", "Domain", k.replace("_", " ").title(), res, SEV["MEDIUM"], v.get("detail", ""))
                    all_test_cases.append(tc_item)
                    tcs_domain.append(tc_item)

                passed = sum(1 for t in tcs_domain if t["Result"] == "PASS")
                failed = len(tcs_domain) - passed
                results["details"]["health_score"] = {
                    "test_cases": tcs_domain,
                    "passed": passed,
                    "failed": failed,
                    "total": len(tcs_domain)
                }

            # Final Site Health Score Calculation (v19 Logic)
            health_meta = calculate_site_health(all_test_cases, results)
            results["health_score"] = health_meta["score"]
            results["health_meta"] = health_meta

            jlog(jid, "="*52, "hdr")
            jlog(jid, f"  COMPLETE — Health Score: {health_meta['score']}/100 ({health_meta['grade']})", "hdr")
            jlog(jid, f"  Total Issues: {tf}  Checks: {tp} Passed / {tf} Failed", "hdr")
            jlog(jid, "="*52, "hdr")

            # ── Full-site image + link crawl ────────────────────────────────
            jlog(jid, "="*52, "hdr")
            jlog(jid, f"  SITE CRAWL — scanning up to {req.max_pages} pages", "hdr")
            jlog(jid, "="*52, "hdr")
            try:
                crawl = await _crawl_site_for_assets(jid, page, req.url, req.max_pages)
                results["image_issues"]         = crawl["image_issues"]
                results["link_issues"]          = crawl["link_issues"]
                results["pages_crawled"]        = crawl["pages_crawled"]
                results["total_images_checked"] = crawl["total_images_checked"]
                results["total_links_checked"]  = crawl["total_links_checked"]
                n_img = len(crawl["image_issues"])
                n_lnk = len(crawl["link_issues"])
                n_pg  = len(crawl["pages_crawled"])
                results["pages_scanned"] = n_pg
                jlog(jid, f"  Crawl done — {n_pg} pages, {n_img} image issues, {n_lnk} broken links", "hdr")
            except Exception as _ce:
                jlog(jid, f"  Crawl warning: {str(_ce)[:80]}", "warn")
                results["image_issues"]  = []
                results["link_issues"]   = []
                results["pages_crawled"] = []

            # ── Location accessibility check (if locations were requested) ──
            if getattr(req, "locations", None):
                jlog(jid, "="*52, "hdr")
                jlog(jid, f"  LOCATION CHECK — {len(req.locations)} region(s)", "hdr")
                jlog(jid, "="*52, "hdr")
                loc_results = await _check_locations_for_qa(
                    jid, req.url, req.locations,
                    use_proxy=req.use_proxy,
                    proxy_session_type=req.proxy_session_type,
                    proxy_protocol=req.proxy_protocol,
                )
                results["location_results"] = loc_results
                accessible = sum(1 for r in loc_results if r["ok"])
                results["location_health"] = round(accessible / len(loc_results) * 100) if loc_results else 0
                jlog(jid, f"  Location check done — {accessible}/{len(loc_results)} accessible ({results['location_health']}%)", "hdr")

            ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
            rp  = REPORTS_DIR/f"qa_{ts2}.json"
            rp.write_text(json.dumps(results, indent=2), encoding='utf-8')
            db.save_report(jid, str(rp), "json")

            # Generate HTML Report
            html_content = _generate_qa_html_report(results)
            hp = REPORTS_DIR/f"qa_{ts2}.html"
            hp.write_text(html_content, encoding='utf-8')
            db.save_report(jid, str(hp), "html")

            # Generate Excel Report
            try:
                xlsx_bytes = _generate_qa_xlsx_report(results)
                if xlsx_bytes:
                    xp = REPORTS_DIR / f"qa_{ts2}.xlsx"
                    xp.write_bytes(xlsx_bytes)
                    results["report_xlsx"] = f"/reports/qa_{ts2}.xlsx"
                    db.save_report(jid, str(xp), "xlsx")
                    jlog(jid, f"✓ Excel report saved: qa_{ts2}.xlsx", "ok")
            except Exception as xlsx_err:
                jlog(jid, f"Warning: Excel report failed: {xlsx_err}", "warn")

            results["report_file"] = str(rp)
            results["report_html"] = f"/reports/qa_{ts2}.html"
            results["report_json"] = f"/reports/qa_{ts2}.json"
            await browser.close()
            jdone(jid, results)
    except Exception as e:
        log.exception("QA Scan"); jerr(jid, str(e))


# ── QA Engine ────────────────────────────────────────────────────────────────

class QAEngine:
    @staticmethod
    async def check_seo(page) -> dict:
        raw = await page.evaluate("""() => {
            const title = document.title || "";
            const desc_el = document.querySelector('meta[name="description"]');
            const h1s = Array.from(document.querySelectorAll("h1")).map(e => e.textContent.trim());
            const canonical = document.querySelector('link[rel="canonical"]');
            const ogTitle = document.querySelector('meta[property="og:title"]');
            const viewport_m = document.querySelector('meta[name="viewport"]');
            const imgCount = document.querySelectorAll("img").length;
            const imgNoAlt = Array.from(document.querySelectorAll("img")).filter(i => !i.alt).length;
            return { title, desc: desc_el ? desc_el.content : "", h1s, canonical: canonical ? canonical.href : null,
                     ogTitle: ogTitle ? ogTitle.content : null, viewportMeta: viewport_m ? viewport_m.content : null,
                     imgCount, imgNoAlt };
        }""")
        tcs = []
        def chk(tc_id, name, ok, sev, detail="", exp="", act=""):
            tcs.append(_tc(tc_id, "SEO", name, "PASS" if ok else "FAIL", sev, detail, exp, act))
        chk("SEO-01", "Title tag present", bool(raw["title"]), SEV["CRITICAL"], raw["title"] or "No title")
        chk("SEO-03", "Meta description present", bool(raw["desc"]), SEV["HIGH"], raw["desc"] or "Missing")
        chk("SEO-05", "Exactly one H1", len(raw["h1s"]) == 1, SEV["HIGH"], f"Found {len(raw['h1s'])} H1(s)")
        chk("SEO-12", "Viewport meta present", bool(raw["viewportMeta"]), SEV["CRITICAL"])
        chk("SEO-14", "All images have alt", raw["imgNoAlt"] == 0, SEV["HIGH"], f"{raw['imgNoAlt']}/{raw['imgCount']} missing alt")
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_accessibility(page) -> dict:
        raw = await page.evaluate("""() => {
            const imgNoAlt = Array.from(document.images).filter(i => !i.hasAttribute("alt")).length;
            const lang = document.documentElement.lang || "";
            const skipLink = !!document.querySelector('a[href="#main"],a[href="#content"],.skip-link');
            const unlabelled = Array.from(document.querySelectorAll('input,select,textarea')).filter(el=>!el.getAttribute('aria-label')&&!el.getAttribute('aria-labelledby')&&!(el.id&&document.querySelector('label[for="'+el.id+'"]'))).length;
            return { imgNoAlt, lang, skipLink, unlabelled, imgCount: document.images.length };
        }""")
        tcs = []
        def chk(tc_id, name, ok, sev, detail="", exp="", act=""):
            tcs.append(_tc(tc_id, "Accessibility", name, "PASS" if ok else "FAIL", sev, detail, exp, act))
        chk("A11Y-01", "HTML lang attribute set", bool(raw["lang"]), SEV["CRITICAL"], f"lang='{raw['lang']}'")
        chk("A11Y-02", "All images have alt text", raw["imgNoAlt"] == 0, SEV["CRITICAL"], f"{raw['imgNoAlt']} missing alt")
        chk("A11Y-03", "All form fields labelled", raw["unlabelled"] == 0, SEV["CRITICAL"], f"{raw['unlabelled']} unlabelled")
        chk("A11Y-04", "Skip navigation link", raw["skipLink"], SEV["HIGH"])
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_performance(page) -> dict:
        raw = await page.evaluate("""() => {
            const n = performance.getEntriesByType('navigation')[0] || {};
            const p = performance.getEntriesByType('paint');
            const totalSize = performance.getEntriesByType('resource').reduce((s, r) => s + (r.transferSize || 0), 0);
            return { ttfb: Math.round(n.responseStart - n.requestStart),
                     fcp: Math.round(p.find(e => e.name === 'first-contentful-paint')?.startTime || 0),
                     load: Math.round(n.loadEventEnd || 0), sizeKB: Math.round(totalSize/1024) };
        }""")
        tcs = []
        def chk(tc_id, name, ok, sev, detail="", exp="", act=""):
            tcs.append(_tc(tc_id, "Performance", name, "PASS" if ok else "FAIL", sev, detail, exp, act))
        chk("PERF-01", "TTFB < 600ms", raw["ttfb"] < 600, SEV["HIGH"], f"{raw['ttfb']}ms")
        chk("PERF-02", "FCP < 1800ms", raw["fcp"] < 1800 if raw["fcp"] else True, SEV["HIGH"], f"{raw['fcp']}ms")
        chk("PERF-04", "Full page load < 4s", raw["load"] < 4000, SEV["HIGH"], f"{raw['load']}ms")
        chk("PERF-05", "Total page weight < 3MB", raw["sizeKB"] < 3000, SEV["MEDIUM"], f"{raw['sizeKB']}KB")
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs), "metrics": raw}

    @staticmethod
    async def check_security(page, response) -> dict:
        headers = {k.lower(): v for k,v in response.headers.items()} if response else {}
        tcs = []
        def chk(tc_id, name, ok, sev, detail="", exp="", act=""):
            tcs.append(_tc(tc_id, "Security", name, "PASS" if ok else "FAIL", sev, detail, exp, act))
        chk("SEC-01", "HTTPS enforced", page.url.startswith("https://"), SEV["CRITICAL"])
        chk("SEC-02", "HSTS header present", "strict-transport-security" in headers, SEV["HIGH"])
        chk("SEC-03", "X-Content-Type-Options", headers.get("x-content-type-options") == "nosniff", SEV["HIGH"])
        chk("SEC-05", "Content-Security-Policy", "content-security-policy" in headers, SEV["HIGH"])
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def build_console_test_cases(errors: list) -> dict:
        tcs = []
        if not errors:
            tcs.append(_tc("CON-01", "Console Errors", "No console errors", "PASS", "PASS"))
            return {"test_cases": tcs, "passed": 1, "failed": 0, "total": 1}
        for idx, raw_msg in enumerate(errors[:10], 1):
            parsed = _describe_console_message(raw_msg)
            tcs.append(_tc(f"CON-{idx:02d}", "Console", parsed["short_title"], "FAIL", parsed["severity"], parsed["detail"]))
        return {"test_cases": tcs, "passed": 0, "failed": len(tcs), "total": len(tcs)}

    @staticmethod
    async def check_broken_links(page, base_url) -> dict:
        import urllib.request, urllib.parse
        broken, ok, tcs = [], 0, []
        try:
            hrefs = await page.evaluate("Array.from(document.querySelectorAll('a[href]')).map(a=>a.href)")
            bd = urllib.parse.urlparse(base_url).netloc
            for h in hrefs[:30]:
                if not h.startswith("http") or urllib.parse.urlparse(h).netloc!=bd: continue
                try:
                    req=urllib.request.Request(h,method="HEAD",headers={"User-Agent":"SiteSentinel/4"})
                    with urllib.request.urlopen(req,timeout=8) as r:
                        if r.status>=400: broken.append({"url":h,"status":r.status})
                        else: ok+=1
                except: broken.append({"url":h,"status":"err"})
        except: pass
        for b in broken: tcs.append(_tc("LINK-ERR", "Links", f"Broken: {b['url'][:50]}", "FAIL", SEV["HIGH"], f"Status: {b['status']}"))
        if ok: tcs.append(_tc("LINK-OK", "Links", f"{ok} links verified", "PASS", "PASS"))
        return {"test_cases": tcs, "passed": ok, "failed": len(broken), "total": ok + len(broken)}

    @staticmethod
    async def check_keyboard_access(page) -> dict:
        raw = await page.evaluate("""() => {
            const focusable = document.querySelectorAll('a[href],button,input,select,textarea,[tabindex]:not([tabindex="-1"])');
            const skipLink = !!document.querySelector('a[href^="#"]');
            return { total: focusable.length, skipLink };
        }""")
        tcs = []
        tcs.append(_tc("KEY-01", "Keyboard", "Focusable elements present", "PASS" if raw["total"]>0 else "FAIL", SEV["CRITICAL"]))
        tcs.append(_tc("KEY-03", "Keyboard", "Skip link present", "PASS" if raw["skipLink"] else "FAIL", SEV["HIGH"]))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_forms(page) -> dict:
        raw = await page.evaluate("""() => {
            return Array.from(document.querySelectorAll("form")).map(f => ({
                hasCSRF: !!f.querySelector('[name*="csrf"],[name*="token"]'),
                submit: f.querySelectorAll('[type="submit"]').length > 0
            }));
        }""")
        tcs = []
        for i, f in enumerate(raw, 1):
            tcs.append(_tc(f"FORM-{i}-CSRF", "Forms", f"Form {i} CSRF", "PASS" if f["hasCSRF"] else "FAIL", SEV["CRITICAL"]))
        if not raw: tcs.append(_tc("FORM-00", "Forms", "No forms found", "PASS", "PASS"))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_typography(page) -> dict:
        raw = await page.evaluate("""() => {
            const fs = parseFloat(window.getComputedStyle(document.body).fontSize);
            return { baseFS: fs };
        }""")
        tcs = []
        tcs.append(_tc("TYPO-01", "Typography", "Base font size ≥ 14px", "PASS" if raw["baseFS"]>=14 else "FAIL", SEV["MEDIUM"], f"{raw['baseFS']}px"))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_html_quality(page) -> dict:
        raw = await page.evaluate("""() => ({
            doctype: !!document.doctype,
            charset: document.characterSet === "UTF-8",
            favicon: !!document.querySelector('link[rel*="icon"]')
        })""")
        tcs = []
        tcs.append(_tc("HTML-01", "HTML Quality", "DOCTYPE declared", "PASS" if raw["doctype"] else "FAIL", SEV["CRITICAL"]))
        tcs.append(_tc("HTML-02", "HTML Quality", "Charset UTF-8", "PASS" if raw["charset"] else "FAIL", SEV["HIGH"]))
        tcs.append(_tc("HTML-03", "HTML Quality", "Favicon present", "PASS" if raw["favicon"] else "FAIL", SEV["MEDIUM"]))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_images(page) -> dict:
        raw = await page.evaluate("""() => {
            return Array.from(document.querySelectorAll("img")).map(img => ({
                src: img.currentSrc || img.src || img.getAttribute('src') || '',
                hasAlt: img.hasAttribute("alt"),
            }));
        }""")
        tcs = []
        page_url = page.url
        missing = [i for i in raw if not i["hasAlt"]]
        missing_srcs = [i["src"] for i in missing if i.get("src")]
        no_alt_count = len(missing)
        detail = f"{no_alt_count}/{len(raw)} images missing alt text"
        tc = _tc("IMG-01", "Images", "All images have alt", "PASS" if no_alt_count == 0 else "FAIL", SEV["HIGH"], detail)
        if no_alt_count > 0:
            tc["fix_hint"] = "Add descriptive alt attributes to all <img> tags. Example: <img src='photo.jpg' alt='A clear description of the image'>"
            tc["missing_alt_images"] = missing_srcs
            tc["page_url"] = page_url
        tcs.append(tc)
        return {"test_cases": tcs, "passed": 1 if no_alt_count == 0 else 0, "failed": no_alt_count, "total": len(raw), "missing_alt_images": missing_srcs}

    @staticmethod
    async def check_navigation(page) -> dict:
        raw = await page.evaluate("""() => ({
            nav: !!document.querySelector("nav"),
            footer: !!document.querySelector("footer")
        })""")
        tcs = []
        tcs.append(_tc("NAV-01", "Navigation", "Nav landmark present", "PASS" if raw["nav"] else "FAIL", SEV["HIGH"]))
        tcs.append(_tc("NAV-04", "Navigation", "Footer landmark present", "PASS" if raw["footer"] else "FAIL", SEV["MEDIUM"]))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

    @staticmethod
    async def check_responsive(page, context, url) -> dict:
        tcs = []
        passed = 0
        for k, vp in VIEWPORTS.items():
            try:
                await page.set_viewport_size({"width": vp["width"], "height": vp["height"]})
                await asyncio.sleep(0.5)
                # Simple check for horizontal scroll
                scroll = await page.evaluate("document.documentElement.scrollWidth > document.documentElement.clientWidth")
                res = "FAIL" if scroll else "PASS"
                tcs.append(_tc(f"RESP-{k}", "Responsive", f"No overflow on {vp['label']}", res, SEV["HIGH"]))
                if res == "PASS": passed += 1
            except:
                tcs.append(_tc(f"RESP-{k}", "Responsive", f"Error testing {vp['label']}", "FAIL", SEV["MEDIUM"]))
        return {"test_cases": tcs, "passed": passed, "failed": len(tcs)-passed, "total": len(tcs)}

    @staticmethod
    async def check_mixed_content(page) -> dict:
        raw = await page.evaluate("""() => {
            const resources = performance.getEntriesByType("resource");
            return resources.filter(r => r.name.startsWith("http:")).map(r => r.name);
        }""")
        tcs = []
        for i, url in enumerate(raw[:5], 1):
            tcs.append(_tc(f"MIXED-{i}", "Security", f"Mixed content: {url[:50]}", "FAIL", SEV["CRITICAL"]))
        if not raw:
            tcs.append(_tc("MIXED-00", "Security", "No mixed content", "PASS", "PASS"))
        return {"test_cases": tcs, "passed": 1 if not raw else 0, "failed": len(tcs) if raw else 0, "total": len(tcs)}

    @staticmethod
    async def check_content_quality(page) -> dict:
        raw = await page.evaluate("""() => {
            const text = document.body.innerText || "";
            const words = text.split(/\\s+/).length;
            const lorum = /lorem ipsum/i.test(text);
            return { words, lorum };
        }""")
        tcs = []
        tcs.append(_tc("CONT-01", "Content", "Word count > 300", "PASS" if raw["words"]>300 else "FAIL", SEV["LOW"], f"{raw['words']} words"))
        tcs.append(_tc("CONT-02", "Content", "No placeholder text", "PASS" if not raw["lorum"] else "FAIL", SEV["MEDIUM"]))
        return {"test_cases": tcs, "passed": sum(1 for t in tcs if t["Result"]=="PASS"), "failed": sum(1 for t in tcs if t["Result"]=="FAIL"), "total": len(tcs)}

async def _run_domain_health(jid: str, url: str, page) -> dict:
    """Site health and domain analysis checks."""
    from urllib.parse import urlparse
    import urllib.request
    parsed = urlparse(url)
    domain = parsed.netloc
    jlog(jid, f"Running: Domain Health Analysis ({domain})", "info")

    health = {"domain": domain, "checks": {}, "score": 0}

    # SSL check
    ssl_ok = url.startswith("https://")
    health["checks"]["ssl"] = {"ok": ssl_ok, "detail": "HTTPS enabled" if ssl_ok else "No HTTPS"}
    jlog(jid, f"  SSL: {'✓' if ssl_ok else '✗'}", "ok" if ssl_ok else "err")

    # Response time
    try:
        t0 = time.time()
        urllib.request.urlopen(url, timeout=10)
        ms = round((time.time()-t0)*1000)
        rt_ok = ms < 2000
        health["checks"]["response_time"] = {"ok": rt_ok, "detail": f"{ms}ms (target <2000ms)", "ms": ms}
        jlog(jid, f"  Response time: {ms}ms", "ok" if rt_ok else "warn")
    except Exception as e:
        health["checks"]["response_time"] = {"ok": False, "detail": str(e)[:50]}

    # Robots.txt
    try:
        robots_url = f"{parsed.scheme}://{domain}/robots.txt"
        r = urllib.request.urlopen(robots_url, timeout=5)
        robots_ok = r.status == 200
        health["checks"]["robots_txt"] = {"ok": robots_ok, "detail": "Found" if robots_ok else "Not found"}
        jlog(jid, f"  robots.txt: {'✓' if robots_ok else '✗'}", "ok" if robots_ok else "warn")
    except:
        health["checks"]["robots_txt"] = {"ok": False, "detail": "Not found or error"}

    # Sitemap.xml
    try:
        sitemap_url = f"{parsed.scheme}://{domain}/sitemap.xml"
        r2 = urllib.request.urlopen(sitemap_url, timeout=5)
        sm_ok = r2.status == 200
        health["checks"]["sitemap"] = {"ok": sm_ok, "detail": "Found" if sm_ok else "Not found"}
        jlog(jid, f"  sitemap.xml: {'✓' if sm_ok else '✗'}", "ok" if sm_ok else "warn")
    except:
        health["checks"]["sitemap"] = {"ok": False, "detail": "Not found"}

    # Favicon
    try:
        fav = await page.evaluate("document.querySelector('link[rel*=\"icon\"]')?.href||''")
        health["checks"]["favicon"] = {"ok": bool(fav), "detail": fav[:60] if fav else "Missing"}
        jlog(jid, f"  Favicon: {'✓' if fav else '✗'}", "ok" if fav else "warn")
    except:
        health["checks"]["favicon"] = {"ok": False, "detail": "Error checking"}

    # Mobile friendly
    try:
        vp = await page.evaluate("document.querySelector('meta[name=\"viewport\"]')?.content||''")
        mob_ok = "width=device-width" in vp
        health["checks"]["mobile_friendly"] = {"ok": mob_ok, "detail": vp[:60] if vp else "Viewport meta missing"}
        jlog(jid, f"  Mobile friendly: {'✓' if mob_ok else '✗'}", "ok" if mob_ok else "warn")
    except:
        health["checks"]["mobile_friendly"] = {"ok": False, "detail": "Error"}

    # Page size
    try:
        size_kb = await page.evaluate("""Math.round(
            performance.getEntriesByType('resource')
            .reduce((s,r)=>s+(r.transferSize||0),0)/1024)""")
        size_ok = size_kb < 3000
        health["checks"]["page_size"] = {"ok": size_ok, "detail": f"{size_kb}KB (target <3000KB)", "kb": size_kb}
        jlog(jid, f"  Page size: {size_kb}KB", "ok" if size_ok else "warn")
    except:
        health["checks"]["page_size"] = {"ok": True, "detail": "Unable to measure"}

    # Score
    passed = sum(1 for v in health["checks"].values() if v.get("ok"))
    total  = len(health["checks"])
    health["score"] = round(passed/total*100) if total else 0
    jlog(jid, f"  Domain health score: {health['score']}/100", "ok")
    return health
