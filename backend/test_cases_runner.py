"""
SiteSentinel — Test cases runner: parse, execute, and report on custom test case sets.
"""
from __future__ import annotations

import asyncio
import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import BROWSER_WS, CONFIG, REPORTS_DIR, log, _run_in_proactor
from core import _open_browser, _resolve_url
from job_manager import jdone, jerr, jlog, jobs
from models import TestCasesRunRequest, LoginCredentialsRequest

# ... all extracted code follows ...


def _parse_test_cases_from_excel(path) -> list:
    """Parse test cases from an .xlsx / .xls file."""
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(path, data_only=True)
        ws  = wb.active
        raw_headers = [str(c.value or '').strip() for c in ws[1]]
        headers_low = [h.lower() for h in raw_headers]

        col_aliases = {
            'id':       ['id', 'test id', 'tc id', 'testid', 'case id', 'no', '#'],
            'name':     ['name', 'test name', 'title', 'test case', 'test case name', 'summary'],
            'type':     ['type', 'test type', 'category', 'test category', 'check type'],
            'page':     ['page', 'url', 'page url', 'path', 'endpoint', 'page/url'],
            'steps':    ['steps', 'description', 'test steps', 'desc', 'action', 'test description', 'scenario'],
            'expected': ['expected', 'expected result', 'expected outcome', 'expected behavior', 'acceptance criteria'],
            'severity': ['severity', 'priority', 'level', 'risk'],
        }
        def find_col(key):
            for alias in col_aliases[key]:
                if alias in headers_low:
                    return headers_low.index(alias)
            return -1
        cols = {k: find_col(k) for k in col_aliases}

        test_cases, counter = [], 1
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v for v in row if v is not None):
                continue
            def _get(k):
                idx = cols[k]
                return str(row[idx] or '').strip() if idx >= 0 and idx < len(row) else ''
            tc = {k: _get(k) for k in col_aliases}
            if not tc['id']:
                tc['id'] = f'TC{counter:03d}'
            if not tc['name']:
                tc['name'] = f'Test Case {counter}'
            if not tc['severity']:
                tc['severity'] = 'Medium'
            counter += 1
            test_cases.append(tc)
        return test_cases
    except Exception as exc:
        log.warning(f"Excel parse error: {exc}")
        return []


def _parse_test_cases_from_pdf(path) -> list:
    """Parse test cases from a PDF file."""
    try:
        import pypdf, re as _re
        reader = pypdf.PdfReader(str(path))
        full_text = '\n'.join(
            page.extract_text() or '' for page in reader.pages
        )
        lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

        # Try structured parsing: look for TC-style headers
        test_cases, current, counter = [], None, 1
        tc_header_re = _re.compile(
            r'^(TC\s*\d+|Test\s+Case\s+\d+|\d{1,3}\s*[.)\-])\s+(.*)', _re.I
        )

        for line in lines:
            m = tc_header_re.match(line)
            if m:
                if current:
                    test_cases.append(current)
                current = {
                    'id':       f'TC{counter:03d}',
                    'name':     m.group(2).strip()[:100],
                    'type':     '',
                    'page':     '',
                    'steps':    '',
                    'expected': '',
                    'severity': 'Medium',
                }
                counter += 1
            elif current:
                llow = line.lower()
                if llow.startswith(('expected:', 'expected result:')):
                    current['expected'] = line.split(':', 1)[-1].strip()[:200]
                elif llow.startswith(('type:', 'category:', 'test type:')):
                    current['type'] = line.split(':', 1)[-1].strip()
                elif llow.startswith(('url:', 'page:', 'path:')):
                    current['page'] = line.split(':', 1)[-1].strip()
                elif llow.startswith(('severity:', 'priority:')):
                    current['severity'] = line.split(':', 1)[-1].strip()
                else:
                    current['steps'] = (current['steps'] + ' ' + line).strip()[:400]

        if current:
            test_cases.append(current)

        # Fallback: split by double newlines when no headers found
        if not test_cases:
            paragraphs = [p.strip() for p in full_text.split('\n\n') if len(p.strip()) > 20]
            for i, para in enumerate(paragraphs[:60], start=1):
                test_cases.append({
                    'id':       f'TC{i:03d}',
                    'name':     para[:80],
                    'type':     '',
                    'page':     '',
                    'steps':    para[:300],
                    'expected': 'Page loads successfully',
                    'severity': 'Medium',
                })

        return test_cases
    except ImportError:
        raise Exception("pypdf not installed — run: pip install pypdf")
    except Exception as exc:
        log.warning(f"PDF parse error: {exc}")
        return []


def _detect_tc_type(tc: dict) -> str:
    """Infer the best execution type from test case fields."""
    combined = ' '.join([
        tc.get('type', ''),
        tc.get('name', ''),
        tc.get('steps', ''),
        tc.get('expected', ''),
    ]).lower()

    if any(k in combined for k in ['seo', 'title tag', 'meta description', 'search engine', 'canonical', 'sitemap']):
        return 'seo'
    if any(k in combined for k in ['alt text', 'aria', 'wcag', 'accessibility', 'screen reader', 'keyboard nav', 'a11y']):
        return 'accessibility'
    if any(k in combined for k in ['form', 'submit', 'input field', 'contact form', 'login form', 'required field', 'field label']):
        return 'form'
    if any(k in combined for k in ['https', 'ssl', 'tls', 'certificate', 'security header', 'hsts', 'csp', 'content security']):
        return 'security'
    if any(k in combined for k in ['broken link', 'dead link', '404', 'link check', 'href']):
        return 'links'
    if any(k in combined for k in ['performance', 'load time', 'page speed', 'lcp', 'cls', 'fid', 'ttfb']):
        return 'performance'
    if any(k in combined for k in ['responsive', 'mobile view', 'tablet', 'viewport', 'media query', 'breakpoint']):
        return 'responsive'
    if any(k in combined for k in ['visual', 'screenshot', 'appearance', 'layout', 'design', 'ui check']):
        return 'visual'
    if any(k in combined for k in ['text content', 'heading', 'paragraph', 'message', 'copy', 'label text', 'display']):
        return 'content'
    return 'navigation'


async def _execute_single_tc(page, tc: dict, base_url: str) -> dict:
    """
    Execute one test case with Playwright.
    Returns a rich result dict containing why_pass / why_fail lists,
    a human-readable reason string, and a screenshot path.
    """
    import time as _time

    tc_page = tc.get('page', '').strip()
    if tc_page.startswith('http'):
        target_url = tc_page
    elif tc_page and tc_page not in ('/', ''):
        target_url = base_url.rstrip('/') + '/' + tc_page.lstrip('/')
    else:
        target_url = base_url

    result: dict = {
        'id':          tc.get('id', '???'),
        'name':        tc.get('name', 'Unnamed'),
        'type':        '',
        'target_url':  target_url,
        'expected':    tc.get('expected', ''),
        'severity':    tc.get('severity', 'Medium'),
        'actual':      '',
        'status':      'fail',   # pass | fail | error
        'reason':      '',       # one-line verdict explanation
        'why_pass':    [],       # ✓ items that contributed to passing
        'why_fail':    [],       # ✗ items that caused failure
        'evidence':    '',       # /reports/tc_ev_*.jpg
        'error':       '',       # set only when status == 'error'
        'duration_ms': 0,
    }

    t0 = _time.time()
    try:
        resp = await page.goto(target_url, timeout=20000, wait_until='domcontentloaded')
        await asyncio.sleep(0.8)

        http_status = resp.status if resp else 0
        title       = (await page.title()) or ''
        tc_type     = _detect_tc_type(tc)
        result['type'] = tc_type

        # ── Screenshot ────────────────────────────────────────
        try:
            safe_id = ''.join(c if c.isalnum() or c in '-_' else '_' for c in str(tc.get('id', 'x')))
            ev_name = f"tc_ev_{safe_id}_{int(_time.time())}.jpg"
            ev_path = REPORTS_DIR / ev_name
            await page.screenshot(
                path=str(ev_path), type='jpeg', quality=80,
                clip={'x': 0, 'y': 0, 'width': 1280, 'height': 900},
            )
            result['evidence'] = f"/reports/{ev_name}"
        except Exception as ss_err:
            log.warning(f"Screenshot failed for {tc.get('id')}: {ss_err}")

        wp = result['why_pass']
        wf = result['why_fail']

        # ── Baseline: HTTP status (every test) ───────────────
        if 200 <= http_status < 400:
            wp.append(f"HTTP {http_status} OK — server responded successfully")
        else:
            wf.append(f"HTTP {http_status} — server returned a non-success status code")

        # ══════════════════════════════════════════════════════
        #  TYPE-SPECIFIC CHECKS
        # ══════════════════════════════════════════════════════

        if tc_type == 'navigation':
            if title.strip():
                wp.append(f'Page title present: "{title[:70]}"')
            else:
                wf.append('Page title is empty — <title> tag missing or blank')

            page_text = await page.evaluate("document.body?.innerText || ''")
            wc = len(page_text.split())
            if wc > 50:
                wp.append(f'Page content loaded ({wc:,} words of visible text)')
            else:
                wf.append(f'Page appears empty or near-empty (only {wc} words of visible text)')

            error_phrases = ['404 not found', 'page not found', '500 internal server', 'access denied', 'forbidden']
            found_errors = [p for p in error_phrases if p in page_text.lower()]
            if found_errors:
                wf.append(f'Error message detected on page: "{found_errors[0]}"')
            else:
                wp.append('No error messages detected in page body')

        elif tc_type == 'seo':
            # Title
            if title.strip():
                tlen = len(title)
                hint = ' ✓ (ideal 30–70 chars)' if 30 <= tlen <= 70 else f' ⚠ ({tlen} chars, ideal 30–70)'
                wp.append(f'Title tag: "{title[:70]}"{hint}')
            else:
                wf.append('Missing <title> tag — required for search engine indexing')

            # H1
            h1s = await page.query_selector_all('h1')
            if len(h1s) == 1:
                h1_text = (await h1s[0].inner_text()).strip()[:60]
                wp.append(f'Single H1 heading: "{h1_text}"')
            elif len(h1s) == 0:
                wf.append('No H1 heading found — every page should have exactly one H1')
            else:
                wf.append(f'{len(h1s)} H1 tags found — should have exactly one H1 per page')

            # Meta description
            md = await page.evaluate(
                "document.querySelector('meta[name=\"description\"]')?.getAttribute('content') || ''"
            )
            if md.strip():
                mlen = len(md)
                if 50 <= mlen <= 160:
                    wp.append(f'Meta description ({mlen} chars, ideal 50–160): "{md[:80]}…"')
                elif mlen < 50:
                    wf.append(f'Meta description too short ({mlen} chars, ideal 50–160): "{md[:50]}"')
                else:
                    wf.append(f'Meta description too long ({mlen} chars, ideal 50–160) — search engines will truncate it')
            else:
                wf.append('Missing meta description — important for search result snippets')

            # Canonical
            canonical = await page.evaluate(
                "document.querySelector('link[rel=\"canonical\"]')?.href || ''"
            )
            if canonical:
                wp.append(f'Canonical URL declared: {canonical[:80]}')

            # Robots
            robots = await page.evaluate(
                "document.querySelector('meta[name=\"robots\"]')?.getAttribute('content') || ''"
            )
            if robots:
                if 'noindex' in robots.lower():
                    wf.append(f'robots meta contains "noindex" — this page is blocked from search indexing')
                else:
                    wp.append(f'robots meta: {robots}')

        elif tc_type == 'accessibility':
            # Image alt text
            imgs = await page.query_selector_all('img')
            if imgs:
                missing_srcs = []
                for img in imgs:
                    alt = await img.get_attribute('alt')
                    if alt is None or alt.strip() == '':
                        src = (await img.get_attribute('src')) or ''
                        missing_srcs.append(src.split('/')[-1][:40] or '(no src)')
                if missing_srcs:
                    wf.append(
                        f'{len(missing_srcs)}/{len(imgs)} images missing alt text: '
                        + ', '.join(missing_srcs[:4])
                        + ('…' if len(missing_srcs) > 4 else '')
                    )
                else:
                    wp.append(f'All {len(imgs)} image(s) have descriptive alt text ✓')
            else:
                wp.append('No images on this page (alt text check not applicable)')

            # Keyboard focus
            focusable = await page.query_selector_all(
                'a[href],button:not([disabled]),input:not([type=hidden]),select,textarea,[tabindex]'
            )
            if focusable:
                wp.append(f'{len(focusable)} keyboard-accessible elements found (links, buttons, inputs)')
            else:
                wf.append('No keyboard-focusable elements found — page may be inaccessible via keyboard')

            # Form labels
            inputs_unlabeled = await page.query_selector_all(
                'input:not([type=hidden]):not([type=submit]):not([type=button]):not([aria-label]):not([title])'
            )
            labeled = await page.query_selector_all('[for],[aria-label],[aria-labelledby]')
            if inputs_unlabeled and not labeled:
                wf.append(f'{len(inputs_unlabeled)} form input(s) may lack accessible labels')
            elif inputs_unlabeled:
                wp.append(f'Form labels or ARIA attributes present ({len(labeled)} labeling elements)')

        elif tc_type == 'form':
            forms   = await page.query_selector_all('form')
            inputs  = await page.query_selector_all('input:not([type=hidden]),textarea,select')
            submits = await page.query_selector_all(
                'button[type=submit],[type=submit],button:not([type=button]):not([type=reset])'
            )

            if forms:
                wp.append(f'{len(forms)} <form> element(s) found on page')
            else:
                wf.append('No <form> element found on this page')

            if inputs:
                field_info = []
                for inp in inputs[:6]:
                    itype = (await inp.get_attribute('type')) or (await inp.evaluate('el=>el.tagName.toLowerCase()'))
                    name  = (await inp.get_attribute('name')) or (await inp.get_attribute('placeholder')) or ''
                    field_info.append(f'{itype}({name[:15]})' if name else itype)
                wp.append(f'{len(inputs)} input field(s): {", ".join(field_info[:6])}')
            else:
                wf.append('No form input fields found')

            if submits:
                btn_text = (await submits[0].inner_text()).strip()[:30]
                wp.append(f'Submit button present: "{btn_text}"')
            else:
                wf.append('No submit button found — form cannot be submitted')

        elif tc_type == 'security':
            if target_url.startswith('https://'):
                wp.append('HTTPS active — connection is encrypted with TLS/SSL')
            else:
                wf.append(
                    'Not using HTTPS — data is transmitted in plaintext; use SSL/TLS certificate'
                )

            if 200 <= http_status < 400:
                wp.append(f'Server reachable: HTTP {http_status}')

            # Check header hints via page evaluate (limited from browser context)
            csp = await page.evaluate(
                "document.querySelector('meta[http-equiv=\"Content-Security-Policy\"]')?.content || ''"
            )
            if csp:
                wp.append('Content-Security-Policy meta tag found')

        elif tc_type == 'links':
            links = await page.query_selector_all('a[href]')
            hrefs = []
            for lnk in links[:30]:
                h = await lnk.get_attribute('href')
                if h: hrefs.append(h)
            internal = sum(1 for h in hrefs if h.startswith('/') or base_url in h)
            external = sum(1 for h in hrefs if h.startswith('http') and base_url not in h)
            broken   = sum(1 for h in hrefs if h in ('#', 'javascript:void(0)', 'javascript:;'))
            wp.append(f'{len(links)} links found: {internal} internal, {external} external')
            if broken:
                wf.append(f'{broken} placeholder/broken href(s) (#, javascript:void) found')
            else:
                wp.append('No obvious placeholder links detected')

        elif tc_type == 'performance':
            timing = await page.evaluate("""
                () => {
                    const t = performance.getEntriesByType('navigation')[0];
                    if (!t) return null;
                    return {
                        ttfb: Math.round(t.responseStart - t.startTime),
                        dom:  Math.round(t.domContentLoadedEventEnd - t.startTime),
                        load: Math.round(t.loadEventEnd - t.startTime),
                    };
                }
            """)
            if timing:
                ttfb = timing['ttfb']; dom = timing['dom']; load = timing['load']
                if ttfb < 600:   wp.append(f'TTFB: {ttfb}ms — good (< 600ms)')
                else:            wf.append(f'TTFB: {ttfb}ms — slow (target < 600ms)')
                if dom < 2000:   wp.append(f'DOM ready: {dom}ms — good (< 2 s)')
                else:            wf.append(f'DOM ready: {dom}ms — slow (target < 2 s)')
                if load < 3000:  wp.append(f'Full page load: {load}ms — good (< 3 s)')
                else:            wf.append(f'Full page load: {load}ms — slow (target < 3 s)')
            else:
                wp.append('Navigation timing unavailable — page loaded without errors')

        elif tc_type == 'content':
            page_text = await page.evaluate("document.body?.innerText || ''")
            exp_text  = tc.get('expected', '').strip()
            if exp_text and len(exp_text) < 200:
                if exp_text.lower() in page_text.lower():
                    wp.append(f'Expected text found verbatim: "{exp_text[:60]}"')
                else:
                    key_words = [w for w in exp_text.split() if len(w) > 4][:6]
                    found_kw  = [w for w in key_words if w.lower() in page_text.lower()]
                    if found_kw and len(found_kw) >= len(key_words) * 0.5:
                        wp.append(
                            f'Key terms matched ({len(found_kw)}/{len(key_words)}): '
                            + ', '.join(f'"{w}"' for w in found_kw[:4])
                        )
                    else:
                        wf.append(
                            f'Expected text not found: "{exp_text[:80]}" '
                            f'(checked {len(key_words)} key terms, found {len(found_kw)})'
                        )
            wc = len(page_text.split())
            if wc > 100:
                wp.append(f'Rich content present: {wc:,} words of visible text')
            elif wc > 20:
                wp.append(f'Some content present: {wc} words')
            else:
                wf.append(f'Very little content on page: only {wc} words')

        else:  # visual / generic
            if title:
                wp.append(f'Page renders with title: "{title[:60]}"')
            else:
                wf.append('Page rendered but no <title> tag found')
            content_size = len(await page.evaluate("document.body?.innerHTML || ''"))
            wp.append(f'Page HTML rendered ({content_size:,} bytes)')

        # ── Verdict ───────────────────────────────────────────
        if wf:
            result['status'] = 'fail'
            result['reason'] = f"FAILED — {wf[0]}" + (f'; and {len(wf)-1} more issue(s)' if len(wf) > 1 else '')
        else:
            result['status'] = 'pass'
            result['reason'] = f"PASSED — all {len(wp)} check(s) succeeded"

        result['actual'] = '; '.join(wp[:2] + [f'✗ {f}' for f in wf[:2]])

    except asyncio.TimeoutError:
        result['status']   = 'error'
        result['error']    = f'Page did not respond within 20 seconds: {target_url}'
        result['actual']   = 'Timeout (> 20 s)'
        result['reason']   = 'ERROR — page load timed out after 20 seconds'
        result['why_fail'] = [
            f'Timeout after 20 seconds — server may be unavailable or the URL is incorrect',
            f'URL attempted: {target_url}',
        ]
    except Exception as exc:
        result['status']   = 'error'
        result['error']    = str(exc)[:200]
        result['actual']   = f'Error: {str(exc)[:80]}'
        result['reason']   = f'ERROR — unexpected exception: {str(exc)[:80]}'
        result['why_fail'] = [f'Playwright exception: {str(exc)[:120]}']

    result['duration_ms'] = round((_time.time() - t0) * 1000)
    return result


# ── HTML Report ───────────────────────────────────────────────────────────────

def _generate_tc_html_report(run_data: dict) -> str:
    from report_engine import ReportBuilder, _score_col, _esc, _sev_badge

    results  = run_data.get('results', [])
    url      = run_data.get('url', '')
    total    = run_data.get('total', 0)
    passed   = run_data.get('passed', 0)
    failed   = run_data.get('failed', 0)
    errors   = run_data.get('errors', 0)
    score    = run_data.get('score', 0)
    ts       = (run_data.get('timestamp', '')[:19].replace('T', ' ')) or datetime.now().strftime('%Y-%m-%d %H:%M')
    viewport = run_data.get('viewport', 'Desktop')

    rb = ReportBuilder(f"Test Case Report — {url}", url, "Test Runner", ts)
    rb.set_score(score, "Pass Rate")
    rb.add_kpi("Total Cases", str(total),  "test cases run",           "#3B82F6")
    rb.add_kpi("Passed",      str(passed), f"{round(passed/max(total,1)*100)}% pass rate", "#22C55E")
    rb.add_kpi("Failed",      str(failed), "assertions failed",         "#EF4444" if failed else "#22C55E")
    rb.add_kpi("Errors",      str(errors), "execution errors",          "#F59E0B" if errors else "#22C55E")
    rb.add_kpi("Viewport",    viewport,    "test environment",          "#8B5CF6")

    type_counts = {}
    for r in results:
        tc_type = r.get('type','generic')
        type_counts[tc_type] = type_counts.get(tc_type,0)+1

    rb.add_charts([
        {"id":"tc_status","title":"Test Status","type":"donut",
         "labels":["Passed","Failed","Errors"],
         "values":[passed, failed, errors],
         "colors":["#22C55E","#EF4444","#F59E0B"]},
        {"id":"tc_types","title":"Tests by Type","type":"bar",
         "labels":list(type_counts.keys()),
         "values":list(type_counts.values()),
         "label":"Count","color":"#3B82F6"},
    ])

    TYPE_COL = {'navigation':'#3B82F6','seo':'#A855F7','accessibility':'#14B8A6',
                'form':'#F59E0B','security':'#EF4444','links':'#6366F1',
                'performance':'#22C55E','content':'#F97316','visual':'#8B5CF6','generic':'#8B949E'}

    rows = ""
    for idx, r in enumerate(results, 1):
        st      = r.get('status','fail')
        tc_type = r.get('type','generic')
        tc_col  = TYPE_COL.get(tc_type,'#8B949E')
        sev     = r.get('severity','Medium')
        border  = {'pass':'#22C55E','fail':'#EF4444','error':'#F59E0B'}.get(st,'#30363D')
        st_col  = {'pass':'#22C55E','fail':'#EF4444','error':'#F59E0B'}.get(st,'#8B949E')
        st_lbl  = {'pass':'✓ PASS','fail':'✗ FAIL','error':'⚠ ERROR'}.get(st,st.upper())

        ev = r.get('evidence','')
        ev_filename = ev.split('/')[-1] if ev else ''
        screenshot_cell = (
            f'<a href="{_esc(ev_filename)}" target="_blank">'
            f'<img src="{_esc(ev_filename)}" alt="Screenshot" style="width:140px;border-radius:4px;border:1px solid #30363D;display:block;">'
            f'<span style="color:#3B82F6;font-size:10px">View →</span></a>'
        ) if ev_filename else '—'

        why_items = r.get('why_pass',[]) + r.get('why_fail',[])
        why_html  = ""
        if why_items:
            col2 = '#22C55E' if r.get('why_pass') else '#EF4444'
            lis  = "".join(f'<li style="margin:3px 0;color:#C9D1D9;font-size:11px">{_esc(str(i))}</li>' for i in why_items[:6])
            why_html = f'<ul style="list-style:none;padding:0;margin:0">{lis}</ul>'
        err = r.get('error','')
        if err and not why_items:
            why_html = f'<span style="color:#F59E0B;font-size:10px;font-family:monospace">{_esc(err[:120])}</span>'

        rows += f"""<tr style="border-left:3px solid {border}">
  <td class="rpt-mono rpt-td-dim" style="font-size:10px">{_esc(r.get('id',''))}<br><span style="color:#3A3A3A">#{idx}</span></td>
  <td><p style="margin:0 0 3px;font-weight:700;color:#F0F0F0;font-size:12px">{_esc(r.get('name',''))}</p>
      <p style="margin:0;font-size:10px;color:#484F58">{_esc(r.get('expected','')[:80])}</p>
      <a class="rpt-link rpt-mono" style="font-size:10px" href="{_esc(r.get('target_url',''))}" target="_blank">{_esc(r.get('target_url','')[:55])}</a></td>
  <td><span style="color:{tc_col};background:{tc_col}1A;border:1px solid {tc_col}33;padding:2px 8px;border-radius:4px;font-size:10px;font-family:monospace;font-weight:700;text-transform:capitalize">{_esc(tc_type)}</span></td>
  <td><span style="color:{st_col};font-weight:700;font-size:12px">{st_lbl}</span><br>
      <span style="font-size:10px;color:{st_col};display:block;margin-top:4px">{_esc(r.get('reason','')[:80])}</span></td>
  <td style="min-width:200px">{why_html}</td>
  <td style="font-size:11px;font-weight:700;color:{'#EF4444' if sev.lower()=='high' else '#F59E0B' if sev.lower()=='medium' else '#22C55E'}">{_esc(sev)}</td>
  <td class="rpt-mono rpt-td-dim" style="font-size:10px">{r.get('duration_ms',0)}ms</td>
  <td>{screenshot_cell}</td>
</tr>"""

    rb.add_section("Test Case Results", "\U0001f52c",
        f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>ID</th><th>Test Name</th><th>Type</th><th>Status</th><th>Detail</th><th>Severity</th><th>Time</th><th>Screenshot</th>
  </tr></thead><tbody>{rows}</tbody></table>
</div></div>""",
        subtitle=f"{passed}/{total} passed · {failed} failed · {errors} errors")

    # Findings for failures
    findings = [{"title": r.get('name',''), "severity": r.get('severity','medium').lower(),
        "category": r.get('type','generic').title(),
        "description": r.get('reason','') or r.get('error','') or 'Check failed',
        "root_cause": "; ".join(str(i) for i in r.get('why_fail',[])) or "See test output.",
        "impact": "Failing test cases may indicate broken functionality for users.",
        "fix": r.get('expected','Verify the expected behavior matches actual behavior.')}
        for r in results if r.get('status') != 'pass']
    if findings:
        rb.add_finding_cards(findings[:20], "Failures & Errors", "⚠️")

    recs = []
    if failed:
        recs.append({"title":f"Fix {failed} failing test(s)","priority":"quick_win",
            "description":"Review each FAIL above and address the root cause.","effort":"Medium","impact":"High"})
    if errors:
        recs.append({"title":f"Resolve {errors} test execution error(s)","priority":"quick_win",
            "description":"Errors indicate infrastructure issues (timeout, network) — check your environment.","effort":"Low","impact":"High"})
    if score < 70:
        recs.append({"title":"Increase pass rate above 70%","priority":"medium",
            "description":f"Current pass rate is {score}%. Focus on the highest-severity failures first.","effort":"Medium","impact":"High"})
    if recs:
        rb.add_recommendations(recs)
    return rb.build()



async def _detect_login_page(page) -> dict:
    """Return {needs_login, form_type, login_url} for the current page."""
    try:
        pwd = await page.query_selector('input[type="password"]')
        if not pwd:
            return {"needs_login": False, "form_type": None, "login_url": page.url}

        has_email = bool(await page.query_selector(
            'input[type="email"], input[name*="email" i], input[placeholder*="email" i], '
            'input[autocomplete="email"], input[autocomplete="username"]'
        ))
        has_user = bool(await page.query_selector(
            'input[name*="username" i], input[name*="user" i], '
            'input[placeholder*="username" i], input[placeholder*="user" i]'
        ))
        form_type = "email" if has_email else ("username" if has_user else "email or username")
        return {"needs_login": True, "form_type": form_type, "login_url": page.url}
    except Exception as e:
        log.warning(f"Login detection error: {e}")
        return {"needs_login": False, "form_type": None, "login_url": getattr(page, "url", "")}


async def _attempt_login(page, username: str, password: str) -> dict:
    """Fill and submit login form. Returns {success, error, redirect_url}."""
    try:
        login_url = page.url

        # Find the primary user/email input (try progressively broader selectors)
        user_field = None
        for sel in [
            'input[type="email"]',
            'input[name*="email" i]',
            'input[autocomplete="email"]',
            'input[autocomplete="username"]',
            'input[name*="username" i]',
            'input[name*="user" i]',
            'input[placeholder*="email" i]',
            'input[placeholder*="username" i]',
            'input[type="text"]',
        ]:
            user_field = await page.query_selector(sel)
            if user_field:
                break

        pwd_field = await page.query_selector('input[type="password"]')

        if not user_field:
            return {"success": False, "error": "Could not find username / email field on this page"}
        if not pwd_field:
            return {"success": False, "error": "Could not find password field on this page"}

        # Fill credentials
        await user_field.triple_click()
        await user_field.fill(username)
        await asyncio.sleep(0.25)
        await pwd_field.triple_click()
        await pwd_field.fill(password)
        await asyncio.sleep(0.25)

        # Find and click submit
        submit = None
        for sel in [
            'button[type="submit"]', 'input[type="submit"]',
            'button:has-text("Login")', 'button:has-text("Sign in")',
            'button:has-text("Log in")', 'button:has-text("Continue")',
            'button:has-text("Submit")',
        ]:
            try:
                submit = await page.query_selector(sel)
                if submit:
                    break
            except Exception:
                pass

        pre_url = page.url
        if submit:
            await submit.click()
        else:
            await pwd_field.press("Enter")

        # Wait for the page to respond
        try:
            await page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            await asyncio.sleep(2)

        await asyncio.sleep(0.5)
        post_url = page.url

        # Still on login page?
        login_indicators = ["/login", "/signin", "/sign-in", "/auth/login", "/wp-login"]
        still_on_login = any(ind in post_url.lower() for ind in login_indicators)
        no_navigation  = (post_url == pre_url)

        if still_on_login or no_navigation:
            page_text = await page.evaluate("document.body?.innerText || ''")
            bad = ["invalid", "incorrect", "wrong", "failed", "error", "try again",
                   "check your", "not found", "no account", "does not exist", "bad credentials"]
            error_msg = "Incorrect credentials — please check your username/email and password"
            for phrase in bad:
                if phrase in page_text.lower():
                    for line in [l.strip() for l in page_text.split("\n") if l.strip()]:
                        if any(p in line.lower() for p in bad) and len(line) < 160:
                            error_msg = line
                            break
                    break
            return {"success": False, "error": error_msg}

        return {"success": True, "error": "", "redirect_url": post_url}
    except Exception as exc:
        return {"success": False, "error": f"Login error: {str(exc)[:120]}"}


async def _run_test_cases_impl(jid: str, req: TestCasesRunRequest):
    import time as _time
    try:
        from playwright.async_api import async_playwright
        tcs = req.test_cases

        jlog(jid, "="*52, "hdr")
        jlog(jid, f"  TEST RUNNER  —  {req.url}", "hdr")
        jlog(jid, f"  {len(tcs)} test case(s) to execute", "hdr")
        jlog(jid, "="*52, "hdr")

        # Pre-check URL
        jlog(jid, "Checking URL accessibility…", "info")
        ok, final_url, st = await _resolve_url(req.url)
        if not ok:
            jerr(jid, f"URL not accessible ({st}): {req.url}")
            return
        if final_url != req.url:
            jlog(jid, f"Resolved: {req.url} → {final_url}", "ok")
        base_url = final_url.rstrip('/')

        VPs = {"desktop":{"width":1920,"height":1080},"mac":{"width":1440,"height":900},
               "laptop":{"width":1366,"height":768},"mobile":{"width":430,"height":932}}
        vp = VPs.get(req.viewport, VPs["desktop"])

        results_list: list = []
        passed = failed = errors = 0

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=True,
                args=["--no-sandbox","--disable-setuid-sandbox",
                      "--disable-blink-features=AutomationControlled"],
            )
            ctx = await browser.new_context(
                viewport=vp,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                ignore_https_errors=True,
            )
            page = await ctx.new_page()

            # ── Login detection ───────────────────────────────────
            jlog(jid, "Checking for login page…", "info")
            try:
                await page.goto(base_url, timeout=15000, wait_until="domcontentloaded")
                await asyncio.sleep(1)
            except Exception as nav_e:
                jlog(jid, f"Initial navigation warning: {nav_e}", "warn")

            login_info = await _detect_login_page(page)

            if login_info["needs_login"]:
                form_type = login_info["form_type"] or "email or username"
                jlog(jid, f"🔐 Login page detected — {form_type} required", "warn")

                # Use credentials if already supplied in the request
                creds: dict | None = None
                if req.login_username and req.login_password:
                    creds = {"username": req.login_username, "password": req.login_password}
                    jlog(jid, f"Using pre-configured credentials for: {req.login_username}", "info")

                if not creds:
                    # Ask frontend for credentials
                    jobs[jid]["partial"] = {
                        "needs_login": True, "login_failed": False,
                        "form_type": form_type, "login_url": login_info["login_url"],
                        "results": [], "passed": 0, "failed": 0, "errors": 0,
                    }
                    jlog(jid, "⏳ Waiting for login credentials (up to 5 min)…", "info")

                    for _ in range(300):          # poll 1 s × 300 = 5 min
                        await asyncio.sleep(1)
                        if jobs[jid].get("cancel"):
                            await browser.close()
                            return
                        creds = jobs[jid].get("login_credentials")
                        if creds:
                            jobs[jid].pop("login_credentials", None)
                            break

                    if not creds or creds.get("username") == "__skip__":
                        jlog(jid, "⚠ Login skipped — running without authentication", "warn")
                        jobs[jid]["partial"]["needs_login"] = False
                        creds = None

                # Attempt login (with up to 3 retries via frontend)
                attempts = 0
                while creds and attempts < 3:
                    attempts += 1
                    jlog(jid, f"🔐 Login attempt {attempts} for {creds['username']}…", "info")
                    result = await _attempt_login(page, creds["username"], creds["password"])

                    if result["success"]:
                        jlog(jid, f"✓ Login successful — authenticated session ready", "ok")
                        jobs[jid]["partial"] = {
                            "needs_login": False, "login_success": True,
                            "login_user": creds["username"],
                            "results": [], "passed": 0, "failed": 0, "errors": 0,
                        }
                        break
                    else:
                        jlog(jid, f"✗ Login failed: {result['error']}", "err")
                        if attempts >= 3:
                            jlog(jid, "⚠ Max login attempts reached — running without authentication", "err")
                            jobs[jid]["partial"] = {
                                "needs_login": False, "login_failed": True,
                                "login_error": result["error"],
                                "results": [], "passed": 0, "failed": 0, "errors": 0,
                            }
                            break

                        # Ask frontend to show error and let user retry
                        jobs[jid]["partial"] = {
                            "needs_login": True, "login_failed": True,
                            "login_error": result["error"],
                            "form_type": form_type, "login_url": login_info["login_url"],
                            "results": [], "passed": 0, "failed": 0, "errors": 0,
                        }
                        jobs[jid].pop("login_credentials", None)

                        # Wait for fresh credentials
                        creds = None
                        for _ in range(300):
                            await asyncio.sleep(1)
                            if jobs[jid].get("cancel"):
                                await browser.close()
                                return
                            creds = jobs[jid].get("login_credentials")
                            if creds:
                                jobs[jid].pop("login_credentials", None)
                                break

                        if not creds or creds.get("username") == "__skip__":
                            jlog(jid, "⚠ Login skipped by user — running without authentication", "warn")
                            jobs[jid]["partial"]["needs_login"] = False
                            break
            else:
                jlog(jid, "✓ No login page detected — proceeding directly", "ok")

            # ── Test case loop ────────────────────────────────────
            for i, tc in enumerate(tcs):
                tc_id   = tc.get('id', f'TC{i+1:03d}')
                tc_name = tc.get('name', f'Test Case {i+1}')
                jlog(jid, f"[{i+1}/{len(tcs)}] {tc_id}: {tc_name}", "info")
                jobs[jid]["progress"] = round(i / len(tcs) * 95)

                res = await _execute_single_tc(page, tc, base_url)

                if res['status'] == 'pass':
                    passed += 1
                    jlog(jid, f"  ✓ PASS  [{res['duration_ms']}ms]  {tc_id}: {res['why_pass'][0][:70] if res['why_pass'] else res['actual'][:60]}", "ok")
                elif res['status'] == 'fail':
                    failed += 1
                    jlog(jid, f"  ✗ FAIL  [{res['duration_ms']}ms]  {tc_id}: {res['why_fail'][0][:70] if res['why_fail'] else res['actual'][:60]}", "err")
                else:
                    errors += 1
                    jlog(jid, f"  ⚠ ERROR [{res['duration_ms']}ms]  {tc_id}: {res['error'][:70]}", "warn")

                results_list.append(res)
                jobs[jid]["partial"] = {
                    "results": results_list.copy(),
                    "passed": passed, "failed": failed, "errors": errors,
                }

            await browser.close()

        total = len(results_list)
        score = round(passed / total * 100) if total else 0
        jlog(jid, "="*52, "hdr")
        jlog(jid, f"  COMPLETE — {passed}/{total} passed  ({score}% pass rate)", "hdr")
        jlog(jid, f"  Failed: {failed}   Errors: {errors}   Total: {total}", "hdr")
        jlog(jid, "="*52, "hdr")

        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        final = {
            "url": base_url, "total": total, "passed": passed,
            "failed": failed, "errors": errors, "score": score,
            "viewport": req.viewport,
            "results": results_list,
            "timestamp": datetime.now().isoformat(),
        }

        # ── JSON report ───────────────────────────────────────
        rp = REPORTS_DIR / f"tc_{ts}.json"
        rp.write_text(json.dumps(final, indent=2), encoding='utf-8')
        final["report_json"] = f"/reports/tc_{ts}.json"
        jlog(jid, f"✓ JSON report saved: tc_{ts}.json", "ok")

        # ── HTML report ───────────────────────────────────────
        try:
            html_content = _generate_tc_html_report(final)
            hp = REPORTS_DIR / f"tc_{ts}.html"
            hp.write_text(html_content, encoding='utf-8')
            final["report_html"] = f"/reports/tc_{ts}.html"
            jlog(jid, f"✓ HTML report saved: tc_{ts}.html", "ok")
        except Exception as html_err:
            jlog(jid, f"Warning: HTML report failed: {html_err}", "warn")

        jobs[jid]["progress"] = 100
        jdone(jid, final)

    except Exception as exc:
        log.exception("Test runner failed")
        jerr(jid, str(exc))


def _run_test_cases(jid: str, req: TestCasesRunRequest):
    """Sync wrapper — delegates to ProactorEventLoop for Playwright on Windows."""
    _run_in_proactor(_run_test_cases_impl(jid, req))
