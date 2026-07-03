"""
SiteSentinel — Pagination / data-integrity report generators.
Run logic (_run_pagination) lives in load_tester.py which imports these.
"""
from __future__ import annotations

from datetime import datetime

from report_utils import _REPORT_CSS, _report_badge, _score_color


def _generate_pagination_html_report(results: dict, url: str = "") -> str:
    from report_engine import ReportBuilder, _score_col, _esc

    pages   = results.get("pages", [])
    checked = results.get("pages_checked", 0)
    records = results.get("records_found", 0)
    dupes   = results.get("duplicates", 0)
    missing = results.get("missing", 0)
    ts      = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    health  = 100 - min(100, dupes * 5 + missing // 10)

    rb = ReportBuilder("Pagination & Data Integrity Report", url, "Pagination Audit", ts)
    rb.set_score(health, "Data Integrity")
    rb.add_kpi("Pages Checked",    str(checked), "pages tested",          "#3B82F6")
    rb.add_kpi("Records Found",    str(records), "total records",         "#22C55E")
    rb.add_kpi("Duplicates",       str(dupes),   "duplicate records",     "#EF4444" if dupes else "#22C55E")
    rb.add_kpi("Missing Records",  str(missing), "records not found",     "#F59E0B" if missing else "#22C55E")

    rb.add_charts([
        {"id":"pag_status","title":"Page Status Distribution","type":"donut",
         "labels":["OK","Warn","Error"],
         "values":[sum(1 for p in pages if p.get("status")=="ok"),
                   sum(1 for p in pages if p.get("status")=="warn"),
                   sum(1 for p in pages if p.get("status")=="err")],
         "colors":["#22C55E","#F59E0B","#EF4444"]},
    ])

    rows = ""
    for p in pages:
        st = p.get("status","ok")
        st_col = "#22C55E" if st=="ok" else "#F59E0B" if st=="warn" else "#EF4444"
        dup_col = "#EF4444" if p.get("duplicates",0) > 0 else "#22C55E"
        srt_col = "#22C55E" if p.get("sort_ok") else "#F59E0B"
        rows += f"""<tr>
  <td class="rpt-td-name">Page {_esc(str(p.get('page','')))} </td>
  <td style="font-family:monospace">{_esc(str(p.get('records','—')))}</td>
  <td style="color:{dup_col};font-weight:700">{p.get('duplicates',0)}</td>
  <td style="color:{srt_col}">{'✓ OK' if p.get('sort_ok') else '⚠ Mismatch'}</td>
  <td style="color:{st_col};font-weight:700">{st.upper()}</td>
  <td class="rpt-td-dim">{_esc(str(p.get('error','') or ''))}</td>
</tr>"""
    rb.add_section("Page-by-Page Results", "\U0001f4c4",
        f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>Page</th><th>Records</th><th>Duplicates</th><th>Sort Order</th><th>Status</th><th>Error</th>
  </tr></thead><tbody>{rows}</tbody></table>
</div></div>""")

    recs = []
    if dupes:
        recs.append({"title":f"Deduplicate {dupes} record(s)","priority":"quick_win",
            "description":"Duplicate records indicate a bug in pagination logic or data source.","effort":"Medium","impact":"High"})
    if missing:
        recs.append({"title":f"Investigate {missing} missing record(s)","priority":"medium",
            "description":"Records expected but not found — check offset/limit logic.","effort":"Medium","impact":"High"})
    if recs:
        rb.add_recommendations(recs)
    return rb.build()



def _generate_pagination_xlsx_report(results: dict, url: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb  = Workbook()
    ws  = wb.active; ws.title = "Pagination Results"
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M")
    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(bold=True, color="C9D1D9", size=10)
    NORM     = Font(size=10)
    thin     = Side(style="thin", color="D0D0D0")
    BDR      = Border(bottom=thin)

    ws.append(["Pagination Test Report"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append(["URL", url]); ws.append(["Generated", ts]); ws.append([])
    for label, val in [("Pages Checked", results.get("pages_checked",0)),
                        ("Records Found", results.get("records_found",0)),
                        ("Duplicates",    results.get("duplicates",0)),
                        ("Missing",       results.get("missing",0))]:
        ws.append([label, val]); ws.cell(ws.max_row,1).font = Font(bold=True,size=10)
    ws.append([])

    headers = ["Page","Records","Duplicates","Sort OK","Status","Error"]
    widths  = [10, 12, 14, 12, 10, 40]
    ws.append(headers)
    for i, w in enumerate(widths, 1):
        c = ws.cell(ws.max_row, i); c.font = HDR_FONT; c.fill = HDR_FILL
        ws.column_dimensions[get_column_letter(i)].width = w

    for p in results.get("pages", []):
        ws.append([p.get("page",""), p.get("records",""), p.get("duplicates",0),
                   "Yes" if p.get("sort_ok") else "No",
                   p.get("status","").upper(), p.get("error","") or ""])
        for i in range(1, 7):
            ws.cell(ws.max_row, i).font = NORM
            ws.cell(ws.max_row, i).border = BDR

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()
