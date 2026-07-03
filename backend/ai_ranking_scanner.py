"""
SiteSentinel — AI Ranking (AIR) scanner: crawl, score, and compare sites for AI discoverability.
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import AI_RANKING_DIR, BROWSER_WS, REPORTS_DIR, log, _run_in_proactor
from core import _open_browser, _url_preflight, _resolve_url
from job_manager import jdone, jerr, jlog, jobs
from models import AIRankingRequest

AI_CRAWLERS = [
    {"name": "GPTBot",           "ua": "GPTBot",           "owner": "OpenAI",      "desc": "ChatGPT training & browsing"},
    {"name": "ClaudeBot",        "ua": "ClaudeBot",        "owner": "Anthropic",   "desc": "Claude AI training"},
    {"name": "PerplexityBot",    "ua": "PerplexityBot",    "owner": "Perplexity",  "desc": "Perplexity AI search"},
    {"name": "Google-Extended",  "ua": "Google-Extended",  "owner": "Google",      "desc": "Gemini / Bard training"},
    {"name": "CCBot",            "ua": "CCBot",            "owner": "CommonCrawl", "desc": "Common Crawl dataset"},
    {"name": "Bytespider",       "ua": "Bytespider",       "owner": "ByteDance",   "desc": "TikTok / Doubao AI"},
    {"name": "FacebookBot",      "ua": "FacebookBot",      "owner": "Meta",        "desc": "Meta AI training"},
    {"name": "Applebot",         "ua": "Applebot",         "owner": "Apple",       "desc": "Apple AI / Siri"},
    {"name": "cohere-ai",        "ua": "cohere-ai",        "owner": "Cohere",      "desc": "Cohere AI training"},
    {"name": "AI2Bot",           "ua": "AI2Bot",           "owner": "Allen AI",    "desc": "AI2 / Dolma dataset"},
    {"name": "anthropic-ai",     "ua": "anthropic-ai",     "owner": "Anthropic",   "desc": "Anthropic research crawl"},
    {"name": "omgili",           "ua": "omgili",           "owner": "Webz.io",     "desc": "AI training data"},
    {"name": "Googlebot",        "ua": "Googlebot",        "owner": "Google",      "desc": "Google Search (SEO reference)"},
]


def _air_fetch(url: str, timeout: int = 10) -> tuple:
    import urllib.request as _ureq, ssl as _ssl
    ctx = _ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl.CERT_NONE
    req = _ureq.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; SiteSentinel/1.0)"})
    try:
        with _ureq.urlopen(req, timeout=timeout, context=ctx) as resp:
            return resp.status, resp.read().decode("utf-8", errors="replace"), dict(resp.headers)
    except Exception as e:
        return 0, str(e), {}


def _air_parse_robots(robots_txt: str) -> dict:
    rules: dict = {}
    current_agents: list = []
    for raw_line in robots_txt.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or ":" not in line:
            continue
        field, _, value = line.partition(":")
        field = field.strip().lower()
        value = value.strip()
        if field == "user-agent":
            current_agents.append(value)
        elif field == "disallow":
            for agent in current_agents:
                rules.setdefault(agent, {"allow": [], "disallow": []})["disallow"].append(value)
        elif field == "allow":
            for agent in current_agents:
                rules.setdefault(agent, {"allow": [], "disallow": []})["allow"].append(value)
        else:
            current_agents = []
    return rules


def _air_crawler_access(rules: dict, crawler_ua: str) -> str:
    def _match(ua_key: str):
        r = rules.get(ua_key)
        if r is None:
            return None
        for p in r.get("disallow", []):
            if p in ("/", "/*"):
                return "blocked"
        return "allowed"
    result = _match(crawler_ua)
    if result:
        return result
    result = _match("*")
    if result:
        return result
    return "unspecified"


def _air_analyze_robots(url: str) -> dict:
    from urllib.parse import urlparse
    parsed = urlparse(url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    status, body, _ = _air_fetch(robots_url)
    if status != 200:
        return {"url": robots_url, "found": False, "raw": "",
                "crawler_access": {c["name"]: "unspecified" for c in AI_CRAWLERS}}
    rules = _air_parse_robots(body)
    access = {c["name"]: _air_crawler_access(rules, c["ua"]) for c in AI_CRAWLERS}
    return {"url": robots_url, "found": True, "raw": body[:3000],
            "crawler_access": access, "rules_count": len(rules)}


async def _air_playwright_data(url: str) -> dict:
    from playwright.async_api import async_playwright
    result = {
        "title": "", "meta_description": "", "h1_count": 0, "h2_count": 0, "h3_count": 0,
        "word_count": 0, "image_count": 0, "images_with_alt": 0, "links_count": 0,
        "canonical": "", "og_tags": {}, "schema_types": [], "has_faq": False,
        "has_howto": False, "has_article": False, "internal_links": 0, "external_links": 0,
        "code_blocks": 0, "tables": 0, "lists": 0, "load_time_ms": 0,
        "https": url.startswith("https://"), "robots_meta": "",
        "structured_data_count": 0, "author": "", "date_published": "",
        "breadcrumb": False, "error": None,
    }
    try:
        async with async_playwright() as pw:
            browser = await _open_browser("air", pw)
            page = await browser.new_page()
            t0 = __import__("time").time()
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            result["load_time_ms"] = round((__import__("time").time() - t0) * 1000)
            result["title"] = (await page.title())[:200]
            result["meta_description"] = (await page.eval_on_selector_all(
                'meta[name="description"]', "els => els.map(e => e.content).join('')")) or ""
            result["canonical"] = (await page.eval_on_selector_all(
                'link[rel="canonical"]', "els => els[0]?.href || ''")) or ""
            result["robots_meta"] = (await page.eval_on_selector_all(
                'meta[name="robots"]', "els => els.map(e => e.content).join(',')")) or ""
            result["h1_count"] = await page.locator("h1").count()
            result["h2_count"] = await page.locator("h2").count()
            result["h3_count"] = await page.locator("h3").count()
            body_text = await page.inner_text("body")
            result["word_count"] = len(body_text.split())
            result["image_count"] = await page.locator("img").count()
            result["images_with_alt"] = len(await page.eval_on_selector_all(
                "img[alt]", "els => els.map(e => e.alt).filter(a => a.trim())"))
            result["code_blocks"] = await page.locator("pre,code").count()
            result["tables"] = await page.locator("table").count()
            result["lists"] = await page.locator("ul,ol").count()
            all_links = await page.eval_on_selector_all("a[href]", f"""els => {{
                const base = new URL("{url}");
                return els.map(e => {{
                    try {{ return new URL(e.href).hostname === base.hostname ? 'internal' : 'external'; }}
                    catch {{ return 'internal'; }}
                }});
            }}""")
            result["internal_links"] = all_links.count("internal")
            result["external_links"] = all_links.count("external")
            result["links_count"] = len(all_links)
            og_raw = await page.eval_on_selector_all(
                'meta[property^="og:"]',
                "els => els.map(e => [e.getAttribute('property'), e.content])")
            result["og_tags"] = dict(og_raw)
            sd_raw = await page.eval_on_selector_all(
                'script[type="application/ld+json"]',
                "els => els.map(e => e.textContent)")
            schema_types = []
            for sd in sd_raw:
                try:
                    obj = json.loads(sd)
                    t = obj.get("@type", "")
                    if isinstance(t, list):
                        schema_types.extend(t)
                    elif t:
                        schema_types.append(t)
                    if t == "FAQPage": result["has_faq"] = True
                    if t == "HowTo": result["has_howto"] = True
                    if t in ("Article", "NewsArticle", "BlogPosting"):
                        result["has_article"] = True
                        if isinstance(obj.get("author"), dict):
                            result["author"] = obj["author"].get("name", "")
                        result["date_published"] = obj.get("datePublished", "")
                    if t == "BreadcrumbList": result["breadcrumb"] = True
                except Exception:
                    pass
            result["schema_types"] = schema_types
            result["structured_data_count"] = len(sd_raw)
            await browser.close()
    except Exception as e:
        result["error"] = str(e)[:200]
    return result


def _air_get_page_data(url: str) -> dict:
    return _run_in_proactor(_air_playwright_data(url))


def _air_score_technical(page: dict, robots: dict) -> tuple:
    checks, score, total = [], 0.0, 0.0
    def chk(name, passed, pts, note=""):
        nonlocal score, total
        checks.append({"name": name, "passed": passed, "pts": pts, "note": note})
        total += pts
        if passed: score += pts
    chk("HTTPS",           page.get("https", False),                     10, "Secure protocol required")
    chk("Canonical URL",   bool(page.get("canonical")),                   8,  "Prevents duplicate content")
    chk("Not noindex",     "noindex" not in page.get("robots_meta","").lower(), 7, "Indexing not blocked")
    chk("Title tag",       bool(page.get("title")),                       8,  "Required for AI citation")
    chk("Load < 3s",       page.get("load_time_ms", 9999) < 3000,        7,  "Fast response for crawlers")
    chk("OG tags",         bool(page.get("og_tags")),                     5,  "Rich preview for AI")
    chk("Meta description",bool(page.get("meta_description")),            5,  "Snippet source for AI")
    return round(score / total * 100) if total else 0, checks


def _air_score_content(page: dict) -> tuple:
    checks, score, total = [], 0.0, 0.0
    def chk(name, passed, pts, note=""):
        nonlocal score, total
        checks.append({"name": name, "passed": passed, "pts": pts, "note": note})
        total += pts
        if passed: score += pts
    wc = page.get("word_count", 0)
    chk("Word count 500+",   wc >= 500,   12, f"Current: {wc} words")
    chk("Word count 1500+",  wc >= 1500,   8, "Comprehensive content preferred by AI")
    chk("H1 tag",            page.get("h1_count", 0) >= 1, 10, "Primary topic signal")
    chk("H2 headings",       page.get("h2_count", 0) >= 2,  8, "Content structure")
    chk("Meta description",  bool(page.get("meta_description")), 8, "AI snippet source")
    chk("Images with alt",   page.get("images_with_alt", 0) > 0, 6, "Accessible content")
    chk("FAQ schema",        page.get("has_faq", False),    10, "High AI citation probability")
    chk("HowTo schema",      page.get("has_howto", False),   8, "Instructional content")
    chk("Article schema",    page.get("has_article", False), 7, "Editorial credibility")
    chk("Lists or tables",   page.get("lists", 0) + page.get("tables", 0) > 1, 5, "Structured content")
    chk("Code examples",     page.get("code_blocks", 0) > 0, 5, "Technical depth")
    return round(score / total * 100) if total else 0, checks


def _air_score_trust(page: dict, url: str) -> tuple:
    checks, score, total = [], 0.0, 0.0
    def chk(name, passed, pts, note=""):
        nonlocal score, total
        checks.append({"name": name, "passed": passed, "pts": pts, "note": note})
        total += pts
        if passed: score += pts
    chk("HTTPS",           page.get("https", False),             12, "Trust signal")
    chk("Author metadata", bool(page.get("author")),             15, "E-E-A-T authorship")
    chk("Date published",  bool(page.get("date_published")),     12, "Content freshness")
    chk("Breadcrumbs",     page.get("breadcrumb", False),        10, "Site structure clarity")
    chk("External links",  page.get("external_links", 0) > 0,    8, "Outbound authority refs")
    chk("OG site name",    bool(page.get("og_tags", {}).get("og:site_name")), 8, "Brand identity")
    return round(score / total * 100) if total else 0, checks


def _air_score_sd(page: dict) -> tuple:
    checks, score, total = [], 0.0, 0.0
    schemas = set(page.get("schema_types", []))
    def chk(name, passed, pts, note=""):
        nonlocal score, total
        checks.append({"name": name, "passed": passed, "pts": pts, "note": note})
        total += pts
        if passed: score += pts
    chk("Any schema",        page.get("structured_data_count", 0) > 0, 20, "Structured data present")
    chk("FAQPage",           "FAQPage" in schemas,     20, "Q&A extraction by AI")
    chk("HowTo",             "HowTo" in schemas,        15, "Instructional schema")
    chk("Article/BlogPost",  bool(schemas & {"Article","NewsArticle","BlogPosting"}), 15, "Editorial schema")
    chk("BreadcrumbList",    "BreadcrumbList" in schemas, 10, "Navigation schema")
    chk("Organization",      "Organization" in schemas, 10, "Entity schema")
    chk("2+ schema types",   len(schemas) >= 2,          10, "Rich schema diversity")
    return round(score / total * 100) if total else 0, checks


def _air_llm_analysis(url: str, scores: dict, page: dict, api_key: str) -> dict:
    import urllib.request as _ureq
    prompt = (
        f"You are an AI visibility expert. Analyze this website audit and provide recommendations.\n"
        f"URL: {url}\n"
        f"Scores: Technical={scores.get('technical',0)}, Content={scores.get('content',0)}, "
        f"Trust={scores.get('trust',0)}, StructuredData={scores.get('structured_data',0)}, Overall={scores.get('overall',0)}\n"
        f"Words: {page.get('word_count',0)}, Schemas: {page.get('schema_types',[])[: 5]}, "
        f"HasFAQ: {page.get('has_faq')}, Author: {bool(page.get('author'))}\n\n"
        f"Respond ONLY with JSON: "
        f'{{\"verdict\":\"...\",\"strengths\":[...],\"critical_issues\":[...],\"quick_wins\":[...],\"roadmap\":[...],\"prompt_examples\":[...]}}'
    )
    payload = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 1024,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()
    req = _ureq.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"}
    )
    try:
        with _ureq.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
            text = data.get("content", [{}])[0].get("text", "{}")
            start, end = text.find("{"), text.rfind("}") + 1
            return json.loads(text[start:end])
    except Exception as e:
        return {"verdict": f"LLM analysis failed: {e}", "strengths": [], "critical_issues": [],
                "quick_wins": [], "roadmap": [], "prompt_examples": []}


async def _run_ai_ranking_impl(jid: str, req: AIRankingRequest):
    jlog(jid, "=" * 56, "hdr")
    jlog(jid, f"  AI VISIBILITY & RANKING AUDIT  —  {req.url}", "hdr")
    jlog(jid, "=" * 56, "hdr")
    if not await _url_preflight(jid, req.url): return
    result: dict = {"url": req.url, "timestamp": datetime.now().isoformat()}

    # Phase 1: robots.txt
    if "robots" in req.checks:
        jlog(jid, "Phase 1 — Analyzing robots.txt…", "info")
        robots = _air_analyze_robots(req.url)
        result["robots"] = robots
        allowed = sum(1 for v in robots.get("crawler_access", {}).values() if v == "allowed")
        blocked = sum(1 for v in robots.get("crawler_access", {}).values() if v == "blocked")
        jlog(jid, f"  robots.txt: {allowed} allowed, {blocked} blocked", "ok" if blocked == 0 else "warn")
    else:
        result["robots"] = {"found": False, "crawler_access": {}}

    jobs[jid]["progress"] = 20

    # Phase 2: Playwright page data
    jlog(jid, "Phase 2 — Extracting on-page signals…", "info")
    page_data = {}
    if any(c in req.checks for c in ["technical", "content", "trust", "structured_data"]):
        try:
            page_data = _air_get_page_data(req.url)
            if page_data.get("error"):
                jlog(jid, f"  Page error: {page_data['error']}", "warn")
            else:
                jlog(jid, f"  {page_data.get('word_count',0)} words, {page_data.get('structured_data_count',0)} schemas", "ok")
        except Exception as e:
            jlog(jid, f"  Playwright failed: {e}", "warn")
            page_data = {"error": str(e), "https": req.url.startswith("https://")}

    jobs[jid]["progress"] = 45

    # Phase 3: Score all dimensions
    jlog(jid, "Phase 3 — Scoring dimensions…", "info")
    scores: dict = {}
    checks_detail: dict = {}
    if "technical" in req.checks:
        sc, ch = _air_score_technical(page_data, result.get("robots", {}))
        scores["technical"] = sc; checks_detail["technical"] = ch
        jlog(jid, f"  Technical: {sc}/100", "ok" if sc >= 70 else "warn")
    if "content" in req.checks:
        sc, ch = _air_score_content(page_data)
        scores["content"] = sc; checks_detail["content"] = ch
        jlog(jid, f"  Content: {sc}/100", "ok" if sc >= 70 else "warn")
    if "trust" in req.checks:
        sc, ch = _air_score_trust(page_data, req.url)
        scores["trust"] = sc; checks_detail["trust"] = ch
        jlog(jid, f"  Trust: {sc}/100", "ok" if sc >= 70 else "warn")
    if "structured_data" in req.checks:
        sc, ch = _air_score_sd(page_data)
        scores["structured_data"] = sc; checks_detail["structured_data"] = ch
        jlog(jid, f"  Structured Data: {sc}/100", "ok" if sc >= 70 else "warn")

    overall = round(
        scores.get("technical", 0) * 0.35 +
        scores.get("content", 0) * 0.40 +
        scores.get("trust", 0) * 0.15 +
        scores.get("structured_data", 0) * 0.10
    )
    scores["overall"] = overall
    result["scores"] = scores
    result["checks"] = checks_detail
    result["page_data"] = page_data
    jobs[jid]["progress"] = 65

    # Phase 4: Competitors
    if "competitors" in req.checks and req.competitor_urls:
        jlog(jid, f"Phase 4 — Analyzing {len(req.competitor_urls)} competitor(s)…", "info")
        competitors = []
        for comp_url in req.competitor_urls[:3]:
            try:
                cp = _air_get_page_data(comp_url)
                cst, _ = _air_score_technical(cp, {})
                csc, _ = _air_score_content(cp)
                ctr, _ = _air_score_trust(cp, comp_url)
                csd, _ = _air_score_sd(cp)
                cs = {"technical": cst, "content": csc, "trust": ctr, "structured_data": csd,
                      "overall": round(cst*0.35 + csc*0.40 + ctr*0.15 + csd*0.10)}
                competitors.append({"url": comp_url, "scores": cs})
                jlog(jid, f"  {comp_url}: overall={cs['overall']}", "ok")
            except Exception as e:
                competitors.append({"url": comp_url, "scores": {}, "error": str(e)})
        result["competitors"] = competitors
    jobs[jid]["progress"] = 80

    # Phase 5: LLM
    if req.use_llm and req.llm_api_key:
        jlog(jid, "Phase 5 — Running Claude AI analysis…", "info")
        result["llm_analysis"] = _air_llm_analysis(req.url, scores, page_data, req.llm_api_key)
        jlog(jid, f"  LLM done.", "ok")
    else:
        result["llm_analysis"] = None
    jobs[jid]["progress"] = 90

    # Save
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    audit_id = f"air_{ts}_{jid[:6]}"
    result["audit_id"] = audit_id
    (AI_RANKING_DIR / f"{audit_id}.json").write_text(json.dumps(result, indent=2), encoding="utf-8")

    try:
        html_content = _generate_air_html_report(result)
        (AI_RANKING_DIR / f"{audit_id}.html").write_text(html_content, encoding="utf-8")
        result["report_html"] = f"/ai-ranking/audit/{audit_id}/html"
        jlog(jid, "✓ HTML report saved", "ok")
    except Exception as e:
        jlog(jid, f"HTML report failed: {e}", "warn")
    try:
        xb = _generate_air_xlsx_report(result)
        if xb:
            (AI_RANKING_DIR / f"{audit_id}.xlsx").write_bytes(xb)
            result["report_xlsx"] = f"/ai-ranking/audit/{audit_id}/xlsx"
            jlog(jid, "✓ Excel report saved", "ok")
    except Exception as e:
        jlog(jid, f"Excel report failed: {e}", "warn")

    # ── Normalise result to flat fields expected by the frontend ─────────────
    from urllib.parse import urlparse as _ulp
    _dom = _ulp(req.url).netloc
    _pd  = result.get("page_data", {})
    _ro  = result.get("robots", {})
    _ca  = _ro.get("crawler_access", {})

    def _air_breakdown(lst):
        bd = {}
        for _i, _c in enumerate(lst):
            _pts = _c.get("pts", 0)
            bd[str(_i)] = {
                "label": _c.get("name", ""),
                "pass":  _c.get("passed", False),
                "pts":   _pts if _c.get("passed") else 0,
                "max":   _pts,
            }
        return bd

    result["domain"]                = _dom
    result["overall_score"]         = scores.get("overall", 0)
    result["technical_score"]       = scores.get("technical", 0)
    result["content_score"]         = scores.get("content", 0)
    result["trust_score"]           = scores.get("trust", 0)
    result["structured_data_score"] = scores.get("structured_data", 0)

    result["robots_analysis"] = {
        "accessible":    _ro.get("found", False),
        "allowed_count": sum(1 for v in _ca.values() if v == "allowed"),
        "blocked_count": sum(1 for v in _ca.values() if v == "blocked"),
        "partial_count": sum(1 for v in _ca.values() if v == "partial"),
        "crawlers": [
            {
                "name":   c["name"],
                "owner":  c["owner"],
                "desc":   c["desc"],
                "status": _ca.get(c["name"], "unspecified"),
                "rule":   _ca.get(c["name"], "unspecified"),
            }
            for c in AI_CRAWLERS
        ],
    }

    result["technical_details"] = {
        "score":          scores.get("technical", 0),
        "breakdown":      _air_breakdown(checks_detail.get("technical", [])),
        "title":          _pd.get("title", ""),
        "is_https":       _pd.get("https", False),
        "load_ms":        _pd.get("load_time_ms", 0),
        "ttfb":           _pd.get("ttfb", 0),
        "canonical":      _pd.get("canonical", ""),
        "og_title":       _pd.get("og_tags", {}).get("og:title", ""),
        "sitemap_found":  _pd.get("sitemap_found", False),
        "internal_links": _pd.get("internal_links", 0),
    }

    result["content_details"] = {
        "score":          scores.get("content", 0),
        "breakdown":      _air_breakdown(checks_detail.get("content", [])),
        "word_count":     _pd.get("word_count", 0),
        "h1s":            [f"H1 #{j+1}" for j in range(_pd.get("h1_count", 0))],
        "h2s":            [f"H2 #{j+1}" for j in range(_pd.get("h2_count", 0))],
        "has_faq":        _pd.get("has_faq", False),
        "has_author":     bool(_pd.get("author", "")),
        "has_date":       bool(_pd.get("date_published", "")),
        "external_links": _pd.get("external_links", 0),
    }

    result["trust_details"] = {
        "score":     scores.get("trust", 0),
        "breakdown": _air_breakdown(checks_detail.get("trust", [])),
    }

    result["structured_data_details"] = {
        "score":         scores.get("structured_data", 0),
        "json_ld_count": _pd.get("structured_data_count", 0),
        "schema_types":  _pd.get("schema_types", []),
        "breakdown":     _air_breakdown(checks_detail.get("structured_data", [])),
    }

    _raw_comps = result.get("competitors", [])
    result["competitors"] = [
        {
            "url":             _c.get("url", ""),
            "overall_score":   _c.get("scores", {}).get("overall", 0),
            "technical_score": _c.get("scores", {}).get("technical", 0),
            "content_score":   _c.get("scores", {}).get("content", 0),
            "trust_score":     _c.get("scores", {}).get("trust", 0),
            "sd_score":        _c.get("scores", {}).get("structured_data", 0),
        }
        for _c in _raw_comps
    ]
    # ─────────────────────────────────────────────────────────────────────────

    jlog(jid, f"✓ AI Ranking complete — overall score: {overall}/100", "ok")
    jdone(jid, result)


def _run_ai_ranking(jid: str, req: AIRankingRequest):
    _run_in_proactor(_run_ai_ranking_impl(jid, req))


_AIR_CSS = """<style>
*{box-sizing:border-box}body{margin:0;font-family:'Inter',system-ui,sans-serif;background:#0D1117;color:#C9D1D9}
.wrap{max-width:1100px;margin:0 auto;padding:32px 24px}
.banner{background:linear-gradient(135deg,#161B22,#1C2128);border:1px solid #30363D;border-radius:20px;padding:32px 36px;margin-bottom:28px;display:flex;align-items:center;gap:32px}
.banner-icon{font-size:40px}.banner-title{font-size:26px;font-weight:800;color:#fff;margin:0 0 4px}
.banner-sub{font-size:13px;color:#8B949E;margin:0}.overall{margin-left:auto;text-align:center}
.overall-val{font-size:52px;font-weight:900}.overall-lbl{font-size:11px;color:#8B949E;text-transform:uppercase;letter-spacing:.06em}
.grid4{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:28px}
.kpi{background:#161B22;border:1px solid #30363D;border-radius:14px;padding:20px;border-top-width:3px}
.kpi-val{font-size:32px;font-weight:800;line-height:1;margin-bottom:4px}.kpi-lbl{font-size:11px;color:#8B949E;text-transform:uppercase;letter-spacing:.05em}
.kpi-weight{font-size:10px;color:#484F58;margin-top:2px}.bar-wrap{height:6px;background:#21262D;border-radius:3px;margin-top:8px}
.bar{height:100%;border-radius:3px}
.card{background:#161B22;border:1px solid #30363D;border-radius:16px;margin-bottom:24px;overflow:hidden}
.card-hdr{padding:14px 20px;border-bottom:1px solid #30363D;display:flex;align-items:center;gap:10px}
.card-hdr-icon{font-size:16px}.card-hdr-title{font-size:14px;font-weight:700;color:#F0F6FC}.card-body{padding:20px}
table{width:100%;border-collapse:collapse}
th{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:#484F58;font-weight:700;padding:10px 14px;text-align:left;border-bottom:1px solid #21262D;background:#0D1117}
td{font-size:12px;padding:9px 14px;border-bottom:1px solid #1C2128;color:#8B949E;vertical-align:top}
td.td-name{color:#E6EDF3;font-weight:600}tr:last-child td{border-bottom:none}
.dot{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:5px}
.badge{display:inline-block;padding:2px 8px;border-radius:6px;font-size:10px;font-weight:700;border:1px solid}
.badge-pass{color:#22C55E;background:rgba(34,197,94,.1);border-color:rgba(34,197,94,.3)}
.badge-fail{color:#EF4444;background:rgba(239,68,68,.1);border-color:rgba(239,68,68,.3)}
.badge-warn{color:#F59E0B;background:rgba(245,158,11,.1);border-color:rgba(245,158,11,.3)}
.check-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px}
.check-item{background:#0D1117;border:1px solid #21262D;border-radius:10px;padding:12px;display:flex;align-items:flex-start;gap:10px}
.pass{color:#22C55E}.fail{color:#EF4444}
.chart-wrap{height:280px;position:relative}
footer{text-align:center;color:#2A2A2A;font-size:11px;margin-top:40px;padding-top:20px;border-top:1px solid #21262D}
</style>"""


def _generate_air_html_report(result: dict) -> str:
    from report_engine import ReportBuilder, _score_col, _esc

    scores      = result.get("scores", {})
    robots      = result.get("robots", {})
    checks      = result.get("checks", {})
    competitors = result.get("competitors", [])
    llm         = result.get("llm_analysis")
    url         = result.get("url", "")
    ts          = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    overall = scores.get("overall", 0)
    dims = [
        ("technical",       "Technical",        "35%"),
        ("content",         "Content Quality",  "40%"),
        ("trust",           "Trust & Auth.",    "15%"),
        ("structured_data", "Structured Data",  "10%"),
    ]

    rb = ReportBuilder("AI Visibility & Ranking Report", url, "AI Ranking Audit", ts)
    rb.set_score(overall, "AI Score")
    for key, label, weight in dims:
        s = scores.get(key, 0)
        rb.add_kpi(label, str(s), f"Weight: {weight}", _score_col(s))

    # Score radar
    rb.add_charts([
        {"id":"air_radar","title":"Dimension Scores","type":"radar",
         "labels":["Technical","Content Quality","Trust","Structured Data"],
         "values":[scores.get(k,0) for k in ["technical","content","trust","structured_data"]],
         "color":"#F5A623"},
        {"id":"air_bar","title":"Score Breakdown","type":"bar",
         "labels":["Technical","Content","Trust","Struct.Data","Overall"],
         "values":[scores.get(k,0) for k in ["technical","content","trust","structured_data","overall"]],
         "label":"Score","color":"#F5A623"},
    ])

    # AI Crawler Access table
    crawler_rows = ""
    for c in AI_CRAWLERS:
        access = robots.get("crawler_access", {}).get(c["name"], "unspecified")
        if access == "allowed":
            badge = '<span style="color:#22C55E;font-weight:700;font-size:11px">ALLOWED</span>'
        elif access == "blocked":
            badge = '<span style="color:#EF4444;font-weight:700;font-size:11px">BLOCKED</span>'
        else:
            badge = '<span style="color:#8B949E;font-weight:700;font-size:11px">UNSPECIFIED</span>'
        crawler_rows += f"""<tr>
  <td class="rpt-td-name">{_esc(c['name'])}</td>
  <td class="rpt-td-dim">{_esc(c['owner'])}</td>
  <td>{badge}</td>
  <td class="rpt-td-dim" style="font-size:10px">{_esc(c['desc'])}</td>
</tr>"""
    rb.add_section("AI Crawler Access", "\U0001f916",
        f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>Crawler</th><th>Owner</th><th>Access</th><th>Description</th>
  </tr></thead><tbody>{crawler_rows}</tbody></table>
</div></div>""")

    # Dimension check sections
    for key, label, _ in dims:
        items = checks.get(key, [])
        if not items:
            continue
        rows = ""
        for c in items:
            icon = "✓" if c.get("passed") else "✗"
            col  = "#22C55E" if c.get("passed") else "#EF4444"
            rows += f"""<div style="display:flex;gap:10px;align-items:flex-start;padding:8px 0;border-bottom:1px solid #21262D">
  <span style="color:{col};font-weight:800;font-size:14px;flex-shrink:0">{icon}</span>
  <div>
    <div style="font-size:12px;font-weight:600;color:#E6EDF3">{_esc(c.get('name',''))}</div>
    <div style="font-size:10px;color:#484F58">{_esc(c.get('note',''))}</div>
  </div>
</div>"""
        passed_n = sum(1 for c in items if c.get("passed"))
        rb.add_section(f"{label} Checks", "⚙",
            f'<div class="rpt-card"><div class="rpt-card-body">{rows}</div></div>',
            subtitle=f"{passed_n}/{len(items)} checks passed")

    # Competitor comparison
    if competitors:
        comp_rows = ""
        for comp in competitors:
            cs = comp.get("scores", {})
            comp_rows += f"""<tr>
  <td class="rpt-mono" style="font-size:11px">{_esc(comp['url'][:55])}</td>
  {"".join(f'<td style="color:{_score_col(cs.get(k,0))};font-weight:700">{cs.get(k,"—")}</td>' for k in ["technical","content","trust","structured_data"])}
  <td style="color:{_score_col(cs.get("overall",0))};font-weight:800;font-size:14px">{cs.get("overall","—")}</td>
</tr>"""
        rb.add_section("Competitor Comparison", "⚔",
            f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>URL</th><th>Technical</th><th>Content</th><th>Trust</th><th>Struct.Data</th><th>Overall</th>
  </tr></thead><tbody>{comp_rows}</tbody></table>
</div></div>""")

    # Claude AI Analysis
    if llm and llm.get("verdict"):
        ai_html = f'<p style="font-size:14px;color:#F0F6FC;line-height:1.6;margin-bottom:16px">{_esc(llm.get("verdict",""))}</p>'
        for section_title, key in [("Strengths","strengths"),("Critical Issues","critical_issues"),
                                    ("Quick Wins","quick_wins"),("90-Day Roadmap","roadmap")]:
            items = llm.get(key, [])
            if items:
                lis = "".join(f"<li style='margin-bottom:4px'>{_esc(str(i))}</li>" for i in items)
                ai_html += f'<p style="font-size:10px;text-transform:uppercase;color:#484F58;margin:12px 0 6px;font-weight:700">{section_title}</p><ul style="margin:0;padding-left:18px;color:#C9D1D9;font-size:13px">{lis}</ul>'
        rb.add_section("Claude AI Analysis", "◎",
            f'<div class="rpt-card"><div class="rpt-card-body">{ai_html}</div></div>')

    # Recommendations from low scores
    recs = []
    for key, label, _ in dims:
        s = scores.get(key, 0)
        if s < 70:
            recs.append({
                "title": f"Improve {label} score ({s}/100)",
                "priority": "quick_win" if s < 50 else "medium",
                "description": f"The {label} dimension scored {s}/100. Review the checks above and address failures.",
                "effort": "Medium", "impact": "High"
            })
    blocked = [c["name"] for c in AI_CRAWLERS
               if robots.get("crawler_access", {}).get(c["name"]) == "blocked"]
    if blocked:
        recs.append({
            "title": f"Unblock AI crawlers: {', '.join(blocked[:3])}",
            "priority": "quick_win",
            "description": "Blocked crawlers cannot index or cite your content in AI responses. Review robots.txt rules.",
            "effort": "Low", "impact": "High"
        })
    if recs:
        rb.add_recommendations(recs)

    rb.add_raw_data({"url": url, "overall_score": overall, "scores": scores})
    return rb.build()


def _generate_air_xlsx_report(result: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""
    wb = Workbook()
    scores = result.get("scores", {})
    robots = result.get("robots", {})
    checks = result.get("checks", {})
    competitors = result.get("competitors", [])
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(bold=True, color="C9D1D9", size=10)
    NORM = Font(size=10)
    thin = Side(style="thin", color="30363D")
    BORDER = Border(bottom=thin)
    CENTER = Alignment(horizontal="center")

    def hdr(ws, cols, widths):
        ws.append(cols)
        for i, (_, w) in enumerate(zip(cols, widths), 1):
            c = ws.cell(ws.max_row, i)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
            ws.column_dimensions[get_column_letter(i)].width = w

    def row(ws, vals):
        ws.append(vals)
        for i in range(1, len(vals)+1):
            ws.cell(ws.max_row, i).font = NORM
            ws.cell(ws.max_row, i).border = BORDER

    ws1 = wb.active; ws1.title = "Executive Summary"
    ws1.append(["AI Visibility & Ranking Report"]); ws1["A1"].font = Font(bold=True, size=14)
    ws1.append(["URL", result.get("url","")]); ws1.append(["Generated", ts]); ws1.append([])
    hdr(ws1, ["Dimension", "Score", "Weight", "Grade"], [24, 10, 10, 12])
    for dim, label, weight in [("technical","Technical","35%"),("content","Content Quality","40%"),
                                ("trust","Trust & Authority","15%"),("structured_data","Structured Data","10%")]:
        s = scores.get(dim, 0)
        row(ws1, [label, s, weight, "A" if s>=90 else "B" if s>=75 else "C" if s>=60 else "D" if s>=45 else "F"])
    ws1.append([]); ws1.append(["Overall", scores.get("overall",0)])

    ws2 = wb.create_sheet("AI Crawler Access")
    hdr(ws2, ["Crawler", "Owner", "Access", "Description"], [18, 16, 14, 36])
    for c in AI_CRAWLERS:
        row(ws2, [c["name"], c["owner"], robots.get("crawler_access",{}).get(c["name"],"unspecified"), c["desc"]])

    for dim_id, dim_label in [("technical","Technical"),("content","Content Quality"),
                               ("trust","Trust & Authority"),("structured_data","Structured Data")]:
        ws = wb.create_sheet(dim_label[:31])
        hdr(ws, ["Check", "Passed", "Points", "Note"], [30, 10, 10, 40])
        for c in checks.get(dim_id, []):
            row(ws, [c.get("name",""), "Yes" if c.get("passed") else "No", c.get("pts",0), c.get("note","")])

    if competitors:
        wsc = wb.create_sheet("Competitor Comparison")
        hdr(wsc, ["URL","Technical","Content","Trust","Struct.Data","Overall"], [40,12,12,12,14,10])
        row(wsc, [result.get("url","")+" (TARGET)", scores.get("technical",""), scores.get("content",""),
                  scores.get("trust",""), scores.get("structured_data",""), scores.get("overall","")])
        for comp in competitors:
            cs = comp.get("scores", {})
            row(wsc, [comp.get("url",""), cs.get("technical",""), cs.get("content",""),
                      cs.get("trust",""), cs.get("structured_data",""), cs.get("overall","")])

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()
