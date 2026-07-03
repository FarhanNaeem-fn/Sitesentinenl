"""
SiteSentinel — Lighthouse / PageSpeed Insights audit scanner.
"""
from __future__ import annotations

import asyncio
import json
import os
import random
from datetime import datetime
from pathlib import Path

from config import REPORTS_DIR, log, _run_in_proactor
from core import _url_preflight
from job_manager import jdone, jlog, jobs
from models import LighthouseRequest
from report_utils import _REPORT_CSS, _report_badge, _score_color


def _generate_lighthouse_html_report(results: dict, url: str = "") -> str:
    from report_engine import ReportBuilder, _score_col, _badge, _sev_badge, _esc

    scores   = results.get("scores", {})
    cwv      = results.get("cwv", {})
    findings = results.get("findings", [])
    ts       = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    CAT_LABELS = {"performance":"Performance","accessibility":"Accessibility",
                  "best-practices":"Best Practices","seo":"SEO","pwa":"PWA"}
    overall = round(sum(v for v in scores.values() if v is not None) /
                    max(len([v for v in scores.values() if v is not None]), 1))

    rb = ReportBuilder("Lighthouse / PageSpeed Audit", url, "Lighthouse", ts)
    rb.set_score(overall, "PSI Score")

    # KPIs
    for k, label in CAT_LABELS.items():
        v = scores.get(k)
        if v is not None:
            rb.add_kpi(label, str(v), "out of 100", _score_col(v))

    # CWV section
    CWV_META = [("lcp","LCP","< 2.5 s"),("cls","CLS","< 0.10"),
                ("fcp","FCP","< 1.8 s"),("si","Speed Index","< 3.4 s"),
                ("tbt","TBT","< 300 ms"),("ttfb","TTFB","< 800 ms")]
    cwv_items = ""
    for k, label, target in CWV_META:
        val = cwv.get(k, "—")
        cwv_items += f"""<div class="rpt-score-item">
  <div class="rpt-score-num" style="color:#3B82F6;font-size:22px">{_esc(str(val))}</div>
  <div class="rpt-score-label">{label}</div>
  <div class="rpt-kpi-sub" style="margin-top:6px">Target: {target}</div>
</div>"""
    rb.add_section("Core Web Vitals", "⚡",
        f'<div class="rpt-scores-grid">{cwv_items}</div>',
        subtitle="Real-world loading, interactivity and visual stability metrics")

    # Score radar
    score_dict = {CAT_LABELS.get(k, k): v for k, v in scores.items() if v is not None}
    rb.add_score_panel(score_dict, radar_id="lh_radar")

    # Charts — score donut + findings severity bar
    sev_counts = {"Pass":0,"Needs Work":0,"Poor":0}
    for f in findings:
        sc = f.get("score")
        if sc is None: continue
        if sc >= 90: sev_counts["Pass"] += 1
        elif sc >= 50: sev_counts["Needs Work"] += 1
        else: sev_counts["Poor"] += 1

    rb.add_charts([
        {"id":"lh_scores","title":"Category Scores","type":"bar",
         "labels": [CAT_LABELS.get(k,k) for k in scores if scores[k] is not None],
         "values": [v for v in scores.values() if v is not None],
         "color":"#3B82F6","label":"Score"},
        {"id":"lh_findings","title":"Findings Breakdown","type":"donut",
         "labels": list(sev_counts.keys()),
         "values": list(sev_counts.values()),
         "colors": ["#22C55E","#F59E0B","#EF4444"]},
    ])

    # Findings as detailed cards
    enriched = []
    for f in findings[:60]:
        sc = f.get("score")
        sev = "pass" if (sc or 0) >= 90 else ("medium" if (sc or 0) >= 50 else "critical")
        enriched.append({
            "title": f.get("title",""),
            "category": f.get("category",""),
            "severity": sev,
            "description": f.get("display_value",""),
            "detail": f"Score: {sc} · Weight: {f.get('weight',0)}",
            "root_cause": ("This audit passed the Lighthouse threshold." if (sc or 0) >= 90
                          else "This audit failed to meet Lighthouse performance standards."),
            "impact": ("Improving this metric will positively affect Core Web Vitals, "
                      "user experience, and Google Search ranking."),
            "fix": f.get("description","See Lighthouse documentation for remediation guidance."),
        })
    rb.add_finding_cards(enriched, "Audit Findings", "🔍")

    # Recommendations
    recs = []
    for f in findings:
        sc = f.get("score")
        if sc is not None and sc < 50:
            recs.append({"title": f.get("title",""), "priority":"quick_win",
                         "description": f"Score {sc}/100 — {f.get('display_value','')}. "
                                        "Address this audit to improve your Lighthouse score.",
                         "effort": "Medium", "impact": "High"})
    if recs:
        rb.add_recommendations(recs[:10])

    rb.add_raw_data({"scores": scores, "cwv": cwv, "findings_count": len(findings)})
    return rb.build()


def _generate_lighthouse_xlsx_report(results: dict, url: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb = Workbook()
    scores   = results.get("scores", {})
    cwv      = results.get("cwv", {})
    findings = results.get("findings", [])
    ts       = datetime.now().strftime("%Y-%m-%d %H:%M")

    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(name="Calibri", bold=True, color="C9D1D9", size=10)
    BOLD     = Font(name="Calibri", bold=True, size=10)
    NORM     = Font(name="Calibri", size=10)
    MONO     = Font(name="Courier New", size=9)
    CENTER   = Alignment(horizontal="center", vertical="center")
    WRAP     = Alignment(wrap_text=True, vertical="top")
    thin     = Side(style="thin", color="30363D")
    BORDER   = Border(bottom=thin)

    def hdr_row(ws, headers, widths):
        ws.append(headers)
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(ws.max_row, i)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
            ws.column_dimensions[get_column_letter(i)].width = w

    def data_row(ws, values, mono_cols=None):
        ws.append(values)
        row = ws.max_row
        for i, v in enumerate(values, 1):
            c = ws.cell(row, i)
            c.font = MONO if (mono_cols and i in mono_cols) else NORM
            c.alignment = WRAP
            c.border = BORDER

    ws1 = wb.active; ws1.title = "Lighthouse Summary"
    ws1.append(["Lighthouse Audit Report"]); ws1["A1"].font = Font(bold=True, size=14)
    ws1.append(["URL", url]); ws1.append(["Generated", ts]); ws1.append([])

    ws1.append(["AUDIT SCORES"])
    ws1.cell(ws1.max_row, 1).font = BOLD
    hdr_row(ws1, ["Category", "Score", "Grade"], [24, 12, 12])
    CAT_LABELS = {"performance":"Performance","accessibility":"Accessibility",
                  "best-practices":"Best Practices","seo":"SEO","pwa":"PWA"}
    for k, label in CAT_LABELS.items():
        sc = scores.get(k)
        grade = "Excellent" if (sc or 0) >= 90 else "Good" if (sc or 0) >= 75 else "Needs Work" if (sc or 0) >= 50 else "Poor"
        data_row(ws1, [label, sc if sc is not None else "N/A", grade])

    ws1.append([])
    ws1.append(["CORE WEB VITALS"])
    ws1.cell(ws1.max_row, 1).font = BOLD
    hdr_row(ws1, ["Metric", "Value", "Target"], [20, 16, 16])
    CWV_META2 = [("lcp","LCP","<2.5s"),("cls","CLS","<0.1"),("fcp","FCP","<1.8s"),
                 ("si","Speed Index","<3.4s"),("tbt","TBT","<300ms"),("ttfb","TTFB","<800ms")]
    for k, label, target in CWV_META2:
        data_row(ws1, [label, str(cwv.get(k, "—")), target])

    ws2 = wb.create_sheet("Audit Findings")
    hdr_row(ws2, ["Category","Audit Title","Score","Display Value","Weight"], [18,50,10,22,10])
    for f in findings:
        data_row(ws2, [f.get("category",""), f.get("title","")[:80],
                       f.get("score","N/A"), str(f.get("display_value",""))[:30],
                       f.get("weight",0)], mono_cols={3,4})

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _run_lighthouse(jid: str, req: LighthouseRequest):
    """Sync wrapper — delegates to ProactorEventLoop for async PSI fetch."""
    _run_in_proactor(_run_lighthouse_impl(jid, req))


async def _run_lighthouse_impl(jid: str, req: LighthouseRequest):
    jlog(jid, "=" * 52, "hdr")
    jlog(jid, f"  LIGHTHOUSE AUDIT  —  {req.url}", "hdr")
    jlog(jid, f"  Device:{req.device}  Mode:{req.browser_mode}  Cats:{','.join(req.categories)}", "hdr")
    jlog(jid, "=" * 52, "hdr")
    if not await _url_preflight(jid, req.url): return

    import aiohttp

    api_key  = os.environ.get("PSI_API_KEY", "")
    strategy = "desktop" if req.device == "desktop" else "mobile"
    cats_str = "&category=".join(req.categories)
    psi_url  = (
        f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
        f"?url={req.url}&strategy={strategy}&category={cats_str}"
    )
    if api_key:
        psi_url += f"&key={api_key}"
        jlog(jid, "Using PSI API key — full quota available.", "info")
    else:
        jlog(jid, "No PSI_API_KEY set — anonymous quota (~1 req/min). "
                  "Add PSI_API_KEY to .env for higher limits.", "warn")

    data = None
    MAX_RETRIES = 3
    for attempt in range(1, MAX_RETRIES + 1):
        jlog(jid, f"Requesting PageSpeed Insights (attempt {attempt}/{MAX_RETRIES})…", "info")
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(psi_url, timeout=aiohttp.ClientTimeout(total=120)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        break
                    elif resp.status == 429:
                        retry_after = int(resp.headers.get("Retry-After", 62))
                        jlog(jid, f"Rate-limited (429). Waiting {retry_after}s before retry…", "warn")
                        if attempt < MAX_RETRIES:
                            await asyncio.sleep(retry_after)
                        else:
                            jlog(jid, "PSI quota exhausted after retries — falling back to simulation.", "warn")
                    else:
                        err_text = await resp.text()
                        jlog(jid, f"PSI API error {resp.status}: {err_text[:200]}", "err")
                        break
        except asyncio.TimeoutError:
            jlog(jid, f"Request timed out (attempt {attempt}).", "err")
        except Exception as e:
            jlog(jid, f"Connection error: {e}", "err")
            break

    if data is None:
        jlog(jid, "⚠ Running in SIMULATION mode — add PSI_API_KEY in .env for real data.", "warn")
        scores = {cat: random.randint(55, 95) for cat in req.categories}
        cwv = {"lcp": "2.1 s", "fid": "18 ms", "cls": "0.08",
               "fcp": "1.6 s", "si": "3.1 s", "tbt": "210 ms", "ttfb": "480 ms"}
        findings = [
            {"category": "performance", "id": "render-blocking-resources",
             "title": "Eliminate render-blocking resources", "score": 45,
             "display_value": "Potential savings of 0.8 s", "weight": 1},
            {"category": "performance", "id": "unused-css-rules",
             "title": "Reduce unused CSS", "score": 60,
             "display_value": "Potential savings of 72 KiB", "weight": 1},
            {"category": "accessibility", "id": "image-alt",
             "title": "Image elements do not have [alt] attributes", "score": 0,
             "display_value": "3 elements", "weight": 3},
            {"category": "seo", "id": "meta-description",
             "title": "Document does not have a meta description", "score": 0,
             "display_value": "", "weight": 1},
            {"category": "best-practices", "id": "uses-https",
             "title": "Uses HTTPS", "score": 100,
             "display_value": "", "weight": 1},
        ]
        jlog(jid, f"Simulation scores: {scores}", "ok")
        jobs[jid]["progress"] = 80

        ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
        res = {"scores": scores, "cwv": cwv, "findings": findings,
               "simulated": True, "url": req.url,
               "simulated_note": "Add PSI_API_KEY in .env for real Lighthouse data."}

        jf = REPORTS_DIR / f"lighthouse_{ts2}.json"
        jf.write_text(json.dumps(res, indent=2), encoding="utf-8")
        res["report_json"] = f"/reports/{jf.name}"
        try:
            hf = REPORTS_DIR / f"lighthouse_{ts2}.html"
            hf.write_text(_generate_lighthouse_html_report(res, req.url), encoding="utf-8")
            res["report_html"] = f"/reports/{hf.name}"
        except Exception as e:
            jlog(jid, f"HTML report failed: {e}", "warn")
        try:
            xb = _generate_lighthouse_xlsx_report(res, req.url)
            if xb:
                xf = REPORTS_DIR / f"lighthouse_{ts2}.xlsx"
                xf.write_bytes(xb)
                res["report_xlsx"] = f"/reports/{xf.name}"
        except Exception as e:
            jlog(jid, f"Excel report failed: {e}", "warn")
        jdone(jid, res)
        return

    jobs[jid]["progress"] = 80

    lighthouse_result = data.get("lighthouseResult", {})
    cats_data = lighthouse_result.get("categories", {})
    aud       = lighthouse_result.get("audits", {})

    scores = {k: round((v.get("score") or 0) * 100) for k, v in cats_data.items()}
    cwv_map = {
        "lcp":  "largest-contentful-paint",
        "fid":  "max-potential-fid",
        "cls":  "cumulative-layout-shift",
        "fcp":  "first-contentful-paint",
        "si":   "speed-index",
        "tbt":  "total-blocking-time",
        "ttfb": "server-response-time",
    }
    cwv = {k: aud.get(v, {}).get("displayValue", "—") for k, v in cwv_map.items()}

    findings = []
    for cat_id, cat in cats_data.items():
        for ref in cat.get("auditRefs", []):
            a   = aud.get(ref["id"], {})
            if not a: continue
            sc2 = a.get("score")
            wt  = ref.get("weight", 0)
            if sc2 is not None and sc2 >= 0.9 and wt == 0: continue
            findings.append({
                "category":      cat_id,
                "id":            ref["id"],
                "title":         a.get("title", "")[:60],
                "score":         round(sc2 * 100) if sc2 is not None else None,
                "display_value": str(a.get("displayValue", ""))[:30],
                "weight":        wt,
            })

    jlog(jid, f"Audit complete! Scores: {scores}", "ok")

    ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
    res = {"scores": scores, "cwv": cwv, "findings": findings, "simulated": False, "url": req.url}

    jf = REPORTS_DIR / f"lighthouse_{ts2}.json"
    jf.write_text(json.dumps(res, indent=2), encoding="utf-8")
    res["report_json"] = f"/reports/{jf.name}"

    try:
        hf = REPORTS_DIR / f"lighthouse_{ts2}.html"
        hf.write_text(_generate_lighthouse_html_report(res, req.url), encoding="utf-8")
        res["report_html"] = f"/reports/{hf.name}"
        jlog(jid, f"✓ HTML report saved: lighthouse_{ts2}.html", "ok")
    except Exception as e:
        jlog(jid, f"HTML report failed: {e}", "warn")

    try:
        xb = _generate_lighthouse_xlsx_report(res, req.url)
        if xb:
            xf = REPORTS_DIR / f"lighthouse_{ts2}.xlsx"
            xf.write_bytes(xb)
            res["report_xlsx"] = f"/reports/{xf.name}"
            jlog(jid, f"✓ Excel report saved: lighthouse_{ts2}.xlsx", "ok")
    except Exception as e:
        jlog(jid, f"Excel report failed: {e}", "warn")

    jdone(jid, res)
