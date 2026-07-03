"""
SiteSentinel — International / localisation report generators.
Run logic (_run_intl, _run_intl_impl) lives in load_tester.py which imports these.
"""
from __future__ import annotations

from datetime import datetime

from report_utils import _REPORT_CSS, _report_badge, _score_color


def _generate_intl_html_report(results: dict, url: str = "") -> str:
    from report_engine import ReportBuilder, _score_col, _esc

    items  = results.get("results", [])
    ts     = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    passed = sum(1 for r in items if r.get("status") == "pass")
    warned = sum(1 for r in items if r.get("status") == "warn")
    failed = sum(1 for r in items if r.get("status") == "err")
    score  = round(passed / len(items) * 100) if items else 0

    rb = ReportBuilder("International & Localisation QA Report", url, "Intl Audit", ts)
    rb.set_score(score, "Pass Rate")
    rb.add_kpi("Regions Tested", str(len(items)), "locales audited",    "#3B82F6")
    rb.add_kpi("Passed",         str(passed),     "locales passing",    "#22C55E")
    rb.add_kpi("Warnings",       str(warned),     "partial issues",     "#F59E0B" if warned else "#22C55E")
    rb.add_kpi("Errors",         str(failed),     "locales with errors","#EF4444" if failed else "#22C55E")

    rb.add_charts([
        {"id":"intl_status","title":"Locale Status Distribution","type":"donut",
         "labels":["Passed","Warnings","Errors"],
         "values":[passed, warned, failed],
         "colors":["#22C55E","#F59E0B","#EF4444"]},
    ])

    rows = ""
    for r in items:
        st     = r.get("status","warn")
        st_col = "#22C55E" if st=="pass" else "#EF4444" if st=="err" else "#F59E0B"
        hl_col = "#22C55E" if r.get("hreflang") else "#EF4444"
        rows += f"""<tr>
  <td class="rpt-td-name">{_esc(r.get('flag',''))} {_esc(r.get('name',''))}</td>
  <td class="rpt-mono rpt-td-dim">{_esc(r.get('locale',''))}</td>
  <td style="font-size:11px;font-weight:600">{_esc(r.get('dir','—'))}</td>
  <td style="color:{hl_col};font-weight:700">{'✓' if r.get('hreflang') else '✗ Missing'}</td>
  <td class="rpt-mono rpt-td-dim">{_esc(str(r.get('charset','—')))}</td>
  <td style="color:{st_col};font-weight:700">{st.upper()}</td>
  <td class="rpt-td-dim">{_esc(str(r.get('error','') or ''))}</td>
</tr>"""
    rb.add_section(f"Results by Region ({len(items)} locales)", "\U0001f5fa️",
        f"""<div class="rpt-card"><div class="rpt-card-body-np rpt-table-wrap">
  <table class="rpt-table"><thead><tr>
    <th>Region</th><th>Locale</th><th>Direction</th><th>hreflang</th><th>Charset</th><th>Status</th><th>Error</th>
  </tr></thead><tbody>{rows}</tbody></table>
</div></div>""")

    # Findings for errors and hreflang missing
    findings = []
    for r in items:
        if r.get("status") == "err":
            findings.append({"title":f"{r.get('flag','')} {r.get('name','')} — Load Error",
                "severity":"high","category":"International",
                "description":str(r.get("error","") or "Region failed to load"),
                "root_cause":"The locale-specific URL returned an error or timed out.",
                "impact":"Users in this region get a broken or missing experience.",
                "fix":"Check server-side locale routing and 5xx errors in your CDN logs."})
        elif not r.get("hreflang"):
            findings.append({"title":f"{r.get('flag','')} {r.get('name','')} — Missing hreflang",
                "severity":"medium","category":"SEO",
                "description":"No hreflang tag found for this locale.",
                "root_cause":"hreflang tags are missing from the page's <head>.",
                "impact":"Search engines may serve the wrong locale to users.",
                "fix":"Add <link rel='alternate' hreflang='xx' href='...'> for each supported locale."})
    if findings:
        rb.add_finding_cards(findings, "Issues Found", "⚠️")
    return rb.build()



def _generate_intl_xlsx_report(results: dict, url: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb  = Workbook()
    ws  = wb.active; ws.title = "International QA"
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M")
    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(bold=True, color="C9D1D9", size=10)
    NORM     = Font(size=10)
    thin     = Side(style="thin", color="D0D0D0")
    BDR      = Border(bottom=thin)

    ws.append(["International QA Report"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append(["URL", url]); ws.append(["Generated", ts]); ws.append([])

    headers = ["Region","Locale","Direction","hreflang","Charset","Status","Error"]
    widths  = [28, 12, 12, 12, 14, 10, 40]
    ws.append(headers)
    for i, w in enumerate(widths, 1):
        c = ws.cell(ws.max_row, i); c.font = HDR_FONT; c.fill = HDR_FILL
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in results.get("results", []):
        row = [f"{r.get('flag','')} {r.get('name','')}".strip(),
               r.get("locale",""), r.get("dir",""), "Yes" if r.get("hreflang") else "No",
               r.get("charset",""), r.get("status","").upper(), r.get("error","") or ""]
        ws.append(row)
        for i in range(1, 8):
            ws.cell(ws.max_row, i).font = NORM
            ws.cell(ws.max_row, i).border = BDR

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()
