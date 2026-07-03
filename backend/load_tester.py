"""
SiteSentinel — load testing, unicorn scenarios, pagination testing, and i18n testing.
"""
from __future__ import annotations

import asyncio
import json
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import REPORTS_DIR, log, _run_in_proactor, _normalize_load_test_type
from job_manager import jdone, jerr, jlog, jobs
from models import LoadTestRequest, UnicornRequest, PaginationRequest, IntlRequest
from intl_scanner import _generate_intl_html_report, _generate_intl_xlsx_report
from pagination_scanner import _generate_pagination_html_report, _generate_pagination_xlsx_report
from core import _url_preflight
from supabase_client import db


def _generate_load_html_report(results: dict) -> str:
    from report_engine import ReportBuilder, _score_col, _esc

    url          = results.get("url", "Unknown")
    ts           = results.get("timestamp", datetime.now().isoformat())[:16].replace("T", " ") + " UTC"
    total_req    = results.get("total_requests", 0)
    total_err    = results.get("total_errors", 0)
    peak_rps     = results.get("peak_rps", 0)
    peak_vu      = results.get("peak_vu", 0)
    p50          = results.get("final_p50", 0)
    p95          = results.get("final_p95", 0)
    p99          = results.get("final_p99", 0)
    avg_ms       = results.get("avg_ms", 0)
    error_rate   = results.get("error_rate", 0)
    test_types   = results.get("test_types", ["load"])

    rps_s   = results.get("rps_series", [])
    p50_s   = results.get("p50_series", [])
    p95_s   = results.get("p95_series", [])
    p99_s   = results.get("p99_series", [])
    err_s   = results.get("error_series", [])
    vu_s    = results.get("vu_series", [])

    # Health score: penalise error rate and high latency
    health = max(0, 100 - int(error_rate * 3) - (20 if p95 > 2000 else 10 if p95 > 1000 else 0))

    rb = ReportBuilder("Load & Performance Test Report", url, "Load Test", ts)
    rb.set_score(health, "Perf Score")

    err_col = "#EF4444" if error_rate > 5 else "#F59E0B" if error_rate > 1 else "#22C55E"
    rb.add_kpi("Total Requests", f"{total_req:,}", f"Peak {peak_rps} req/s", "#3B82F6")
    rb.add_kpi("Total Errors",   f"{total_err:,}", f"{error_rate}% error rate", err_col)
    rb.add_kpi("Peak VUs",       str(peak_vu),    "concurrent virtual users", "#A855F7")
    rb.add_kpi("P50 Latency",    f"{p50} ms",     "median response time",
               "#22C55E" if p50 < 500 else "#F59E0B" if p50 < 1500 else "#EF4444")
    rb.add_kpi("P95 Latency",    f"{p95} ms",     "95th percentile",
               "#22C55E" if p95 < 1000 else "#F59E0B" if p95 < 3000 else "#EF4444")
    rb.add_kpi("P99 Latency",    f"{p99} ms",     "99th percentile tail",
               "#22C55E" if p99 < 2000 else "#F59E0B" if p99 < 5000 else "#EF4444")

    # Time-series charts
    labels = [str(i + 1) for i in range(len(rps_s))]
    rb.add_charts([
        {"id": "ld_rps_vu", "title": "Throughput & Virtual Users Over Time", "type": "line",
         "labels": labels,
         "datasets": [
             {"label": "RPS",  "data": rps_s, "color": "#3B82F6", "fill": False},
             {"label": "VUs",  "data": vu_s,  "color": "#A855F7", "fill": False},
         ]},
        {"id": "ld_latency", "title": "Latency Percentiles Over Time (ms)", "type": "line",
         "labels": labels,
         "datasets": [
             {"label": "P50", "data": p50_s, "color": "#22C55E", "fill": False},
             {"label": "P95", "data": p95_s, "color": "#F59E0B", "fill": False},
             {"label": "P99", "data": p99_s, "color": "#EF4444", "fill": False},
         ]},
    ])

    # Step-by-step table
    rows = ""
    for i, (rps, p50v, p95v, p99v, err, vu) in enumerate(zip(rps_s, p50_s, p95_s, p99_s, err_s, vu_s), 1):
        ec = "#EF4444" if err > 5 else "#F59E0B" if err > 1 else "#22C55E"
        rows += f"""<tr>
  <td class="rpt-td-mono">{i}</td>
  <td>{vu}</td>
  <td style="color:#3B82F6;font-weight:700">{rps}</td>
  <td style="color:#22C55E">{p50v} ms</td>
  <td style="color:#F59E0B">{p95v} ms</td>
  <td style="color:#EF4444">{p99v} ms</td>
  <td style="color:{ec};font-weight:700">{err}%</td>
</tr>"""

    rb.add_section("Step-by-Step Metrics", "📋",
        f"""<div class="rpt-card">
  <div class="rpt-card-body-np rpt-table-wrap">
    <table class="rpt-table">
      <thead><tr>
        <th>Step</th><th>VUs</th><th>RPS</th>
        <th>P50</th><th>P95</th><th>P99</th><th>Error %</th>
      </tr></thead>
      <tbody>{rows or "<tr><td colspan='7' style='text-align:center;padding:24px;color:var(--text3)'>No step data</td></tr>"}</tbody>
    </table>
  </div>
</div>""",
        subtitle=f"Test types: {', '.join(str(t).upper() for t in test_types)}")

    # Recommendations
    recs = []
    if error_rate > 5:
        recs.append({"title":"High Error Rate Detected","priority":"quick_win",
            "description":f"Error rate of {error_rate}% exceeds the 5% threshold. "
                          "Check server capacity, connection pool limits, and rate limiting.",
            "effort":"Medium","impact":"Critical"})
    if p95 > 2000:
        recs.append({"title":"P95 Latency Exceeds 2 s","priority":"quick_win",
            "description":f"95th percentile latency is {p95} ms. "
                          "Profile slow endpoints, add caching, and optimise database queries.",
            "effort":"Medium","impact":"High"})
    if p99 > 5000:
        recs.append({"title":"P99 Tail Latency Critical","priority":"medium",
            "description":f"1 in 100 requests takes {p99} ms. "
                          "Investigate GC pauses, lock contention, or long-running queries.",
            "effort":"High","impact":"High"})
    if error_rate < 1 and p95 < 1000:
        recs.append({"title":"Performance Budget Maintained","priority":"longterm",
            "description":"Results are within healthy thresholds. "
                          "Continue monitoring to catch regressions before they reach users.",
            "effort":"Low","impact":"Ongoing"})
    if recs:
        rb.add_recommendations(recs)

    rb.add_raw_data({"url": url, "total_requests": total_req, "peak_rps": peak_rps,
                     "peak_vu": peak_vu, "p50": p50, "p95": p95, "p99": p99,
                     "error_rate": error_rate, "test_types": test_types})
    return rb.build()

async def _run_load_test(jid: str, req: LoadTestRequest):
    selected_types = req.test_types or [req.test_type]
    selected_types = [_normalize_load_test_type(item) for item in selected_types]
    selected_types = [t for i, t in enumerate(selected_types) if t not in selected_types[:i]]
    if not selected_types:
        selected_types = ["load"]

    jlog(jid, "=" * 52, "hdr")
    jlog(jid, f"  LOAD TEST  —  {req.url}", "hdr")
    jlog(jid, f"  Test types: {', '.join(selected_types).upper()}", "hdr")
    jlog(jid, f"  VUs:{req.virtual_users}  Dur:{req.duration_min}min  Ramp:{req.ramp_up_sec}s", "hdr")
    jlog(jid, "=" * 52, "hdr")
    if not await _url_preflight(jid, req.url): return

    try:
        import aiohttp as _aiohttp
    except ImportError:
        jlog(jid, "ERROR: aiohttp not installed", "err")
        jerr(jid, "aiohttp library required for load testing")
        return

    all_results = []
    combined = {
        "rps_series": [],
        "p50_series": [],
        "p95_series": [],
        "p99_series": [],
        "error_series": [],
        "vu_series": [],
    }
    total_requests = 0
    total_errors = 0
    peak_rps = 0
    peak_vu = 0

    for idx, test_type in enumerate(selected_types, start=1):
        if jobs[jid].get("cancel"):
            jlog(jid, "Load test cancelled by user", "warn")
            break

        jlog(jid, f"  [{idx}/{len(selected_types)}] Running {test_type.upper()}", "info")
        metrics = await _run_load_test_type(jid, req, test_type, _aiohttp)
        if not metrics:
            continue

        all_results.append(metrics)
        combined["rps_series"].extend(metrics["rps_series"])
        combined["p50_series"].extend(metrics["p50_series"])
        combined["p95_series"].extend(metrics["p95_series"])
        combined["p99_series"].extend(metrics["p99_series"])
        combined["error_series"].extend(metrics["error_series"])
        combined["vu_series"].extend(metrics["vu_series"])
        total_requests += metrics.get("total_requests", 0)
        total_errors += metrics.get("total_errors", 0)
        peak_rps = max(peak_rps, metrics.get("peak_rps", 0))
        peak_vu = max(peak_vu, metrics.get("peak_vu", 0))
        jobs[jid]["partial"] = {
            "current_vu": metrics["vu_series"][-1] if metrics["vu_series"] else 0,
            "current_rps": metrics["rps_series"][-1] if metrics["rps_series"] else 0,
            "current_p50": metrics["p50_series"][-1] if metrics["p50_series"] else 0,
            "current_p95": metrics["p95_series"][-1] if metrics["p95_series"] else 0,
            "current_p99": metrics["p99_series"][-1] if metrics["p99_series"] else 0,
            "current_err": metrics["error_series"][-1] if metrics["error_series"] else 0,
            "current_type": test_type,
        }

    if not all_results:
        jerr(jid, "No load test results were collected")
        return

    average_ms = round(sum(run.get("avg_ms", 0) * run.get("total_requests", 0) for run in all_results) / max(total_requests, 1))
    combined_metrics = {
        **combined,
        "url": req.url,
        "timestamp": datetime.now().isoformat(),
        "test_types": selected_types,
        "total_requests": total_requests,
        "total_errors": total_errors,
        "peak_rps": peak_rps,
        "peak_vu": peak_vu,
        "final_p50": all_results[-1].get("final_p50", 0),
        "final_p95": all_results[-1].get("final_p95", 0),
        "final_p99": all_results[-1].get("final_p99", 0),
        "avg_ms": average_ms,
        "error_rate": round((total_errors / max(total_requests, 1)) * 100, 2),
        "runs": all_results,
    }

    try:
        ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
        rp = REPORTS_DIR / f"load_{ts2}.json"
        hp = REPORTS_DIR / f"load_{ts2}.html"
        rp.write_text(json.dumps(combined_metrics, indent=2), encoding='utf-8')
        hp.write_text(_generate_load_html_report(combined_metrics), encoding='utf-8')
        db.save_report(jid, str(rp), "json")
        db.save_report(jid, str(hp), "html")
        combined_metrics["report_json"] = f"/reports/{rp.name}"
        combined_metrics["report_html"] = f"/reports/{hp.name}"
        # Excel report
        xlsx_bytes = _generate_load_xlsx_report(combined_metrics)
        if xlsx_bytes:
            xp = REPORTS_DIR / f"load_{ts2}.xlsx"
            xp.write_bytes(xlsx_bytes)
            combined_metrics["report_xlsx"] = f"/reports/{xp.name}"
            db.save_report(jid, str(xp), "xlsx")
            jlog(jid, f"✓ Excel report saved: load_{ts2}.xlsx", "ok")
    except Exception as e:
        jlog(jid, f"Warning: failed to write load report files: {e}", "warn")

    jlog(jid, "=" * 52, "hdr")
    jlog(jid, f"  COMPLETE  Peak VUs:{peak_vu}  Peak RPS:{peak_rps}  Errors:{combined_metrics['error_rate']:.1f}%", "hdr")
    jlog(jid, f"  Total Requests:{total_requests}  P95 Latency:{combined_metrics['final_p95']}ms", "hdr")
    jlog(jid, "=" * 52, "hdr")
    jdone(jid, combined_metrics)


async def _run_load_test_type(jid: str, req: LoadTestRequest, test_type: str, aiohttp_module) -> Optional[Dict[str, Any]]:
    url = req.url
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    duration_sec = req.duration_min * 60
    steps = max(req.duration_min * 2, 4)
    step_duration = duration_sec / steps
    response_times = []
    step_metrics = []
    total_requests = 0
    total_errors = 0
    timeout_config = aiohttp_module.ClientTimeout(total=req.timeout_sec)

    async with aiohttp_module.ClientSession(timeout=timeout_config) as session:
        for step_idx in range(steps):
            if jobs[jid].get("cancel"):
                jlog(jid, "Test cancelled by user", "warn")
                break

            ratio = (step_idx / max(steps - 1, 1)) if steps > 1 else 1.0
            if test_type == "spike":
                vu = int(req.virtual_users * 10) if 0.35 < ratio < 0.65 else int(req.virtual_users * ratio * 0.5 + 5)
            elif test_type == "stress":
                vu = int(req.virtual_users * (1 + ratio * 0.6))
            elif test_type == "breakpoint":
                vu = int(req.virtual_users * ratio * 1.8)
            elif test_type == "endurance":
                vu = int(req.virtual_users * (0.6 + ratio * 0.4))
            else:
                ramp_steps = max(1, req.ramp_up_sec // 15)
                vu = int(req.virtual_users * min(1.0, ratio * steps / ramp_steps))
            vu = max(1, vu)

            step_start = time.time()
            step_errors = 0
            step_times = []
            tasks = [_make_load_request(session, url) for _ in range(vu)]
            results = await asyncio.gather(*tasks, return_exceptions=True)

            for result in results:
                if isinstance(result, tuple) and len(result) == 2:
                    response_ms, success = result
                    response_times.append(response_ms)
                    step_times.append(response_ms)
                    if not success:
                        step_errors += 1
                else:
                    step_errors += 1
                    step_times.append(req.timeout_sec * 1000)

            total_requests += len(results)
            total_errors += step_errors

            if step_times:
                step_times.sort()
                n = len(step_times)
                p50 = step_times[max(0, int(n * 0.50) - 1)]
                p95 = step_times[max(0, int(n * 0.95) - 1)]
                p99 = step_times[max(0, min(int(n * 0.99), n - 1))]
                rps = round(len(results) / max(time.time() - step_start, 0.001))
                err_pct = round((step_errors / max(len(results), 1)) * 100, 2)
            else:
                p50 = p95 = p99 = 0
                rps = 0
                err_pct = 100.0

            step_metrics.append({
                "rps": rps,
                "p50": p50,
                "p95": p95,
                "p99": p99,
                "error_pct": err_pct,
                "vu": vu,
            })
            jobs[jid]["progress"] = int((step_idx + 1) / steps * 95)
            lv = "err" if err_pct > 5 else "warn" if err_pct > 1 else "ok"
            jlog(jid, f"{test_type.upper()} | VUs:{vu:4d} | RPS:{rps:5d} | P50:{p50:4d}ms | P95:{p95:4d}ms | Errors:{err_pct:.1f}%", lv)
            jobs[jid]["partial"] = {
                "current_vu": vu,
                "current_rps": rps,
                "current_p50": p50,
                "current_p95": p95,
                "current_p99": p99,
                "current_err": err_pct,
                "current_type": test_type,
            }

            elapsed = time.time() - step_start
            if elapsed < step_duration:
                await asyncio.sleep(step_duration - elapsed)
            if req.think_time_ms > 0 and step_idx < steps - 1:
                await asyncio.sleep(min(req.think_time_ms / 1000.0, 0.5))

    if not response_times:
        return None

    response_times.sort()
    n = len(response_times)
    final_p50 = response_times[max(0, int(n * 0.50) - 1)]
    final_p95 = response_times[max(0, int(n * 0.95) - 1)]
    final_p99 = response_times[max(0, min(int(n * 0.99), n - 1))]
    avg_ms = round(sum(response_times) / n)
    return {
        "test_type": test_type,
        "rps_series": [m["rps"] for m in step_metrics],
        "p50_series": [m["p50"] for m in step_metrics],
        "p95_series": [m["p95"] for m in step_metrics],
        "p99_series": [m["p99"] for m in step_metrics],
        "error_series": [m["error_pct"] for m in step_metrics],
        "vu_series": [m["vu"] for m in step_metrics],
        "total_requests": total_requests,
        "total_errors": total_errors,
        "peak_rps": max((m["rps"] for m in step_metrics), default=0),
        "peak_vu": max((m["vu"] for m in step_metrics), default=0),
        "final_p50": final_p50,
        "final_p95": final_p95,
        "final_p99": final_p99,
        "avg_ms": avg_ms,
    }



async def _make_load_request(session, url: str) -> tuple:
    """Make a single load test request and return (response_ms, success)."""
    try:
        start = time.time()
        async with session.get(url, ssl=False) as resp:
            await resp.read()
            ms = round((time.time() - start) * 1000)
            success = 200 <= resp.status < 400
            return (ms, success)
    except asyncio.TimeoutError:
        ms = round((time.time() - start) * 1000)
        return (ms, False)
    except Exception as e:
        ms = round((time.time() - start) * 1000)
        return (ms, False)


# ── Unicorn Suite ────────────────────────────────────────────────────────────

UNICORN_SCENARIOS = {
    "login_browse_checkout": [
        ("Navigate to homepage",     "GET",  "/"),
        ("Load login page",          "GET",  "/login"),
        ("Submit credentials",       "POST", "/api/auth/login"),
        ("Browse product catalogue", "GET",  "/products"),
        ("View product detail",      "GET",  "/products/1"),
        ("Add to cart",              "POST", "/api/cart/add"),
        ("Load checkout",            "GET",  "/checkout"),
        ("Submit order",             "POST", "/api/orders"),
        ("Confirmation page",        "GET",  "/order/confirm"),
    ],
    "search_filter_view": [
        ("Homepage",                 "GET",  "/"),
        ("Search query",             "GET",  "/search?q=test"),
        ("Apply filter",             "GET",  "/search?q=test&category=1"),
        ("View result detail",       "GET",  "/item/42"),
        ("Related items",            "GET",  "/api/related/42"),
    ],
    "api_auth_crud": [
        ("Health check",             "GET",  "/api/health"),
        ("Authenticate",             "POST", "/api/auth/token"),
        ("Create resource",          "POST", "/api/resources"),
        ("Read resource",            "GET",  "/api/resources/1"),
        ("Update resource",          "PUT",  "/api/resources/1"),
        ("List resources",           "GET",  "/api/resources"),
        ("Delete resource",          "DELETE","/api/resources/1"),
    ],
    "homepage_nav_form": [
        ("Homepage",                 "GET",  "/"),
        ("About page",               "GET",  "/about"),
        ("Contact page",             "GET",  "/contact"),
        ("Submit contact form",      "POST", "/contact"),
        ("Thank you page",           "GET",  "/thank-you"),
    ],
}

async def _run_unicorn(jid: str, req: UnicornRequest):
    import urllib.request as _ur, urllib.parse as _up
    jlog(jid,"="*52,"hdr")
    jlog(jid,f"  🦄 UNICORN SUITE  —  {req.url}","hdr")
    jlog(jid,f"  Scenario: {req.scenario}  VUs: {req.virtual_users}","hdr")
    jlog(jid,"="*52,"hdr")
    if not await _url_preflight(jid, req.url): return

    steps = UNICORN_SCENARIOS.get(req.scenario, UNICORN_SCENARIOS["login_browse_checkout"])
    parsed = _up.urlparse(req.url)
    base   = f"{parsed.scheme}://{parsed.netloc}"

    results = []
    total_p=0; total_f=0

    for iteration in range(min(req.virtual_users, 5)):
        jlog(jid, f"\n--- Virtual User #{iteration+1} ---", "hdr")
        for step_name, method, path in steps:
            if jobs[jid].get("cancel"): break
            full_url = base + path
            t0 = time.time()
            try:
                hdrs = {"User-Agent": "SiteSentinel-Unicorn/3.0", **req.headers}
                r = _ur.Request(full_url, method=method, headers=hdrs)
                with _ur.urlopen(r, timeout=10) as resp:
                    ms  = round((time.time()-t0)*1000)
                    ok  = resp.status < 400
                    status = resp.status
            except Exception as e:
                ms     = round((time.time()-t0)*1000)
                ok     = False
                status = 0
                err_msg = str(e)[:60]

            results.append({"step":step_name,"method":method,"url":full_url,
                            "status":status,"ms":ms,"ok":ok})
            if ok: total_p+=1; jlog(jid,f"  ✓ {step_name}  [{method}] {status}  {ms}ms","ok")
            else:  total_f+=1; jlog(jid,f"  ✗ {step_name}  [{method}] {status}  {ms}ms","err")
            await asyncio.sleep(.05)

        jobs[jid]["progress"] = int((iteration+1)/min(req.virtual_users,5)*90)

    health = round(total_p/(total_p+total_f)*100) if (total_p+total_f)>0 else 0
    jlog(jid,"","info")
    jlog(jid,"="*52,"hdr")
    jlog(jid,f"  UNICORN COMPLETE  {total_p} passed  {total_f} failed  {health}% success","hdr")
    jdone(jid,{"scenario":req.scenario,"steps":results,"passed":total_p,
               "failed":total_f,"health":health,"virtual_users":req.virtual_users})


# ── Pagination ────────────────────────────────────────────────────────────────

async def _run_pagination(jid: str, req: PaginationRequest):
    import urllib.request as _ur, json as _j
    jlog(jid,f"Pagination test: {req.url}","hdr")
    _check_url = req.url.replace("{page}", "1").replace("{size}", "10")
    if not await _url_preflight(jid, _check_url): return
    total_pages=(req.total_records+req.per_page-1)//req.per_page
    seen=set(); dupes=0; missing=0; records=0; pages_r=[]

    for pg in range(1,min(total_pages,15)+1):
        if jobs[jid].get("cancel"): break
        try:
            url=req.url.replace("{page}",str(pg)).replace("{size}",str(req.per_page))
            if "{page}" not in req.url:
                sep="&" if "?" in url else "?"
                url=f"{url}{sep}page={pg}&per_page={req.per_page}"
            r=_ur.urlopen(url,timeout=10)
            data=_j.loads(r.read())
            items=data if isinstance(data,list) else data.get("data",data.get("items",data.get("results",[])))
            ids=[str(item.get(req.id_field,i)) for i,item in enumerate(items)]
            dps=[id_ for id_ in ids if id_ in seen]; dupes+=len(dps); seen.update(ids); records+=len(items)
            sort_ok=True
            pages_r.append({"page":pg,"records":len(items),"duplicates":len(dps),"sort_ok":sort_ok,
                            "status":"ok" if len(dps)==0 else "warn"})
            jlog(jid,f"Page {pg}: {len(items)} records  dups={len(dps)}","ok" if len(dps)==0 else "warn")
        except Exception as e:
            pages_r.append({"page":pg,"error":str(e)[:60],"status":"err"})
            jlog(jid,f"Page {pg}: ERROR — {e}","err")
        jobs[jid]["progress"]=int(pg/min(total_pages,15)*95)
        await asyncio.sleep(.1)

    missing=max(0,req.total_records-records)
    jlog(jid,f"COMPLETE: {len(pages_r)} pages · {records} records · {dupes} dups · {missing} missing",
         "ok" if dupes==0 and missing==0 else "warn")
    pag_res = {"pages_checked":len(pages_r),"records_found":records,
               "duplicates":dupes,"missing":missing,"pages":pages_r,"url":req.url}
    ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        hf = REPORTS_DIR / f"pagination_{ts2}.html"
        hf.write_text(_generate_pagination_html_report(pag_res, req.url), encoding='utf-8')
        pag_res["report_html"] = f"/reports/{hf.name}"
        jlog(jid, f"✓ HTML report saved: pagination_{ts2}.html", "ok")
    except Exception as e:
        jlog(jid, f"Warning: Pagination HTML report failed: {e}", "warn")
    try:
        xlsx_bytes = _generate_pagination_xlsx_report(pag_res, req.url)
        if xlsx_bytes:
            xf = REPORTS_DIR / f"pagination_{ts2}.xlsx"
            xf.write_bytes(xlsx_bytes)
            pag_res["report_xlsx"] = f"/reports/{xf.name}"
            jlog(jid, f"✓ Excel report saved: pagination_{ts2}.xlsx", "ok")
    except Exception as e:
        jlog(jid, f"Warning: Pagination Excel report failed: {e}", "warn")
    jdone(jid, pag_res)


# ── International ─────────────────────────────────────────────────────────────

def _run_intl(jid: str, req: IntlRequest):
    """Sync wrapper — delegates to ProactorEventLoop for Playwright."""
    _run_in_proactor(_run_intl_impl(jid, req))

async def _run_intl_impl(jid: str, req: IntlRequest):
    from playwright.async_api import async_playwright
    LOCALE_META={
        "en-GB":{"name":"United Kingdom","flag":"🇬🇧","dir":"LTR"},
        "en-US":{"name":"United States","flag":"🇺🇸","dir":"LTR"},
        "de-DE":{"name":"Germany","flag":"🇩🇪","dir":"LTR"},
        "fr-FR":{"name":"France","flag":"🇫🇷","dir":"LTR"},
        "ar-AE":{"name":"UAE (Arabic)","flag":"🇦🇪","dir":"RTL"},
        "ur-PK":{"name":"Pakistan (Urdu)","flag":"🇵🇰","dir":"RTL"},
        "fa-IR":{"name":"Iran (Persian)","flag":"🇮🇷","dir":"RTL"},
        "ja-JP":{"name":"Japan","flag":"🇯🇵","dir":"LTR"},
        "zh-CN":{"name":"China","flag":"🇨🇳","dir":"LTR"},
        "hi-IN":{"name":"India (Hindi)","flag":"🇮🇳","dir":"LTR"},
        "pt-BR":{"name":"Brazil","flag":"🇧🇷","dir":"LTR"},
        "es-ES":{"name":"Spain","flag":"🇪🇸","dir":"LTR"},
        "ko-KR":{"name":"South Korea","flag":"🇰🇷","dir":"LTR"},
        "ru-RU":{"name":"Russia","flag":"🇷🇺","dir":"LTR"},
        "tr-TR":{"name":"Turkey","flag":"🇹🇷","dir":"LTR"},
        "nl-NL":{"name":"Netherlands","flag":"🇳🇱","dir":"LTR"},
    }
    results=[]
    jlog(jid,f"International QA: {req.url}  ({len(req.locales)} locales)","hdr")
    if not await _url_preflight(jid, req.url): return
    async with async_playwright() as pw:
        for i,locale in enumerate(req.locales):
            if jobs[jid].get("cancel"): break
            meta=LOCALE_META.get(locale,{"name":locale,"flag":"🌐","dir":"LTR"})
            jlog(jid,f"Testing {meta['flag']} {meta['name']} ({locale})","info")
            try:
                br=await pw.chromium.launch(headless=True)
                cx=await br.new_context(locale=locale,extra_http_headers={"Accept-Language":locale})
                pg=await cx.new_page()
                await pg.goto(req.url,timeout=20000,wait_until="domcontentloaded")
                hreflang = await pg.evaluate(f"""() => {{
                    const target = "{locale}".toLowerCase();
                    const langOnly = target.split('-')[0];
                    const links = Array.from(document.querySelectorAll('link[rel="alternate"][hreflang]'));
                    return links.some(l => {{
                        const val = l.hreflang.toLowerCase();
                        return val === target || val === langOnly || val.startsWith(langOnly + "-");
                    }});
                }}""")
                charset =await pg.evaluate("document.characterSet||document.charset||'unknown'")
                dir_a   =await pg.evaluate("document.documentElement.dir||document.body?.dir||''")
                exp_dir =meta["dir"].lower()
                dir_ok  =(exp_dir=="ltr" or (exp_dir=="rtl" and dir_a.lower()=="rtl"))
                status  ="pass" if hreflang and charset.upper() in ("UTF-8","UTF8") else "warn"
                results.append({"locale":locale,"name":meta["name"],"flag":meta["flag"],
                                 "dir":meta["dir"],"hreflang":hreflang,"charset":charset,
                                 "dir_ok":dir_ok,"status":status})
                jlog(jid,f"  hreflang={hreflang}  charset={charset}  dir_ok={dir_ok}",
                     "ok" if status=="pass" else "warn")
                await br.close()
            except Exception as e:
                results.append({"locale":locale,"name":meta["name"],"flag":meta["flag"],
                                 "error":str(e)[:60],"status":"err"})
                jlog(jid,f"  ERROR: {e}","err")
            jobs[jid]["progress"]=int((i+1)/len(req.locales)*95)
            await asyncio.sleep(.2)
    jlog(jid,f"COMPLETE: {len(results)} regions","hdr")
    intl_res = {"results": results, "url": req.url}
    ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        hf = REPORTS_DIR / f"intl_{ts2}.html"
        hf.write_text(_generate_intl_html_report(intl_res, req.url), encoding='utf-8')
        intl_res["report_html"] = f"/reports/{hf.name}"
        jlog(jid, f"✓ HTML report saved: intl_{ts2}.html", "ok")
    except Exception as e:
        jlog(jid, f"Warning: Intl HTML report failed: {e}", "warn")
    try:
        xlsx_bytes = _generate_intl_xlsx_report(intl_res, req.url)
        if xlsx_bytes:
            xf = REPORTS_DIR / f"intl_{ts2}.xlsx"
            xf.write_bytes(xlsx_bytes)
            intl_res["report_xlsx"] = f"/reports/{xf.name}"
            jlog(jid, f"✓ Excel report saved: intl_{ts2}.xlsx", "ok")
    except Exception as e:
        jlog(jid, f"Warning: Intl Excel report failed: {e}", "warn")
    jdone(jid, intl_res)


# ── Load Test Excel Report ────────────────────────────────────────────────────

def _generate_load_xlsx_report(results: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb  = Workbook()
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M")
    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(bold=True, color="C9D1D9", size=10)
    NORM     = Font(size=10)
    thin     = Side(style="thin", color="D0D0D0")
    BDR      = Border(bottom=thin)

    ws1 = wb.active; ws1.title = "Summary"
    ws1.append(["Load Test Report"]); ws1["A1"].font = Font(bold=True, size=14)
    ws1.append(["URL", results.get("url","")]); ws1.append(["Generated", ts]); ws1.append([])

    metrics = [
        ("Total Requests",  results.get("total_requests",0)),
        ("Total Errors",    results.get("total_errors",0)),
        ("Peak RPS",        results.get("peak_rps",0)),
        ("Peak Virtual Users", results.get("peak_vu",0)),
        ("Avg Response (ms)", results.get("avg_ms",0)),
        ("P50 Latency (ms)",  results.get("final_p50",0)),
        ("P95 Latency (ms)",  results.get("final_p95",0)),
        ("P99 Latency (ms)",  results.get("final_p99",0)),
        ("Error Rate (%)",    results.get("error_rate",0)),
    ]
    ws1.append(["Metric","Value"])
    ws1.cell(ws1.max_row,1).font = HDR_FONT; ws1.cell(ws1.max_row,2).font = HDR_FONT
    ws1.cell(ws1.max_row,1).fill = HDR_FILL; ws1.cell(ws1.max_row,2).fill = HDR_FILL
    for label, val in metrics:
        ws1.append([label, val])
        ws1.cell(ws1.max_row,1).font = Font(bold=True,size=10)
        ws1.cell(ws1.max_row,2).font = NORM

    ws2 = wb.create_sheet("Time Series")
    hdrs = ["Second","Virtual Users","RPS","P50 (ms)","P95 (ms)","P99 (ms)","Error Rate (%)"]
    widths = [10, 16, 10, 12, 12, 12, 16]
    ws2.append(hdrs)
    for i, w in enumerate(widths, 1):
        c = ws2.cell(1, i); c.font = HDR_FONT; c.fill = HDR_FILL
        ws2.column_dimensions[get_column_letter(i)].width = w

    rps_s = results.get("rps_series",[])
    p50_s = results.get("p50_series",[])
    p95_s = results.get("p95_series",[])
    p99_s = results.get("p99_series",[])
    err_s = results.get("error_series",[])
    vu_s  = results.get("vu_series",[])
    for idx in range(len(rps_s)):
        ws2.append([idx+1, vu_s[idx] if idx<len(vu_s) else 0,
                    rps_s[idx], p50_s[idx] if idx<len(p50_s) else 0,
                    p95_s[idx] if idx<len(p95_s) else 0,
                    p99_s[idx] if idx<len(p99_s) else 0,
                    err_s[idx] if idx<len(err_s) else 0])
        for i in range(1, 8):
            ws2.cell(ws2.max_row, i).font = NORM
            ws2.cell(ws2.max_row, i).border = BDR

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue()
